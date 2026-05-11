# sgd/pts/pts_export.py
from __future__ import annotations

import io

from flask import send_file, abort

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .helpers import (
    _val,
    ensure_pts_schema,
    require_clinica_id,
    fetch_pts_by_id,
    fetch_paciente_full,
    fetch_participantes,
    registrar_log,
)


# ============================================================
# TIMBRE · HELPERS
# ============================================================

def _row_dict(row, cur=None):
    if not row:
        return {}

    if isinstance(row, dict) or hasattr(row, "keys"):
        return dict(row)

    if cur and cur.description:
        cols = [d[0] for d in cur.description]
        return dict(zip(cols, row))

    return {}


def _img_bytes(v):
    if not v:
        return None
    if isinstance(v, memoryview):
        return v.tobytes()
    if isinstance(v, bytearray):
        return bytes(v)
    if isinstance(v, bytes):
        return v
    return None


def _buscar_timbre_config(conn, clinica_id: int) -> dict:
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT *
            FROM clinica_configuracoes
            WHERE clinica_id = %s
            ORDER BY id DESC
            LIMIT 1
        """, (int(clinica_id),))

        return _row_dict(cur.fetchone(), cur)

    except Exception:
        return {}

    finally:
        try:
            cur.close()
        except Exception:
            pass


def _draw_img_fit(c, img_bytes, x, y, box_w, box_h, preserve=True, opacity=None):
    if not img_bytes:
        return False

    try:
        from reportlab.lib.utils import ImageReader

        bio = io.BytesIO(img_bytes)
        reader = ImageReader(bio)

        if opacity is not None:
            try:
                c.saveState()
                c.setFillAlpha(opacity)
                c.setStrokeAlpha(opacity)
            except Exception:
                opacity = None

        iw, ih = reader.getSize()

        if preserve and iw and ih:
            scale = min(box_w / iw, box_h / ih)
            draw_w = iw * scale
            draw_h = ih * scale
            draw_x = x + (box_w - draw_w) / 2
            draw_y = y + (box_h - draw_h) / 2
        else:
            draw_x, draw_y, draw_w, draw_h = x, y, box_w, box_h

        c.drawImage(
            reader,
            draw_x,
            draw_y,
            width=draw_w,
            height=draw_h,
            mask="auto",
            preserveAspectRatio=True,
            anchor="c",
        )

        if opacity is not None:
            try:
                c.restoreState()
            except Exception:
                pass

        return True

    except Exception:
        try:
            if opacity is not None:
                c.restoreState()
        except Exception:
            pass
        return False


def _draw_timbre_page(c, w, h, timbre: dict):
    """
    Desenha o timbre da clínica em cada página do PDF.

    Usa:
    - cabecalho_img_bin
    - logo_bin como marca d'água
    - rodape_texto
    - rodape_img_bin / rodape_img_2_bin / rodape_img_3_bin
    - cor_listra_topo
    """

    from reportlab.lib.units import cm
    from reportlab.lib.colors import HexColor

    # Listra superior
    cor = (timbre.get("cor_listra_topo") or "#0f766e").strip()

    if cor and cor.lower() != "transparent":
        try:
            c.setFillColor(HexColor(cor))
        except Exception:
            c.setFillColor(HexColor("#0f766e"))

        c.rect(0, h - 10, w, 10, fill=1, stroke=0)

    # Cabeçalho
    cabecalho = _img_bytes(timbre.get("cabecalho_img_bin"))
    if cabecalho:
        _draw_img_fit(
            c,
            cabecalho,
            x=2 * cm,
            y=h - 4.2 * cm,
            box_w=w - 4 * cm,
            box_h=2.4 * cm,
        )

    # Marca d'água
    logo = _img_bytes(timbre.get("logo_bin"))
    if logo:
        _draw_img_fit(
            c,
            logo,
            x=(w - 11 * cm) / 2,
            y=(h - 11 * cm) / 2,
            box_w=11 * cm,
            box_h=11 * cm,
            opacity=0.10,
        )

    # Rodapé texto
    rodape_texto = (timbre.get("rodape_texto") or "").strip()
    footer_y = 1.55 * cm

    if rodape_texto:
        c.setFont("Helvetica", 7.5)
        c.setFillColor(HexColor("#334155"))

        linhas = rodape_texto.splitlines()[:4]
        y = footer_y + 1.05 * cm

        for linha in linhas:
            c.drawString(2 * cm, y, linha[:125])
            y -= 9

    # Rodapés imagem
    rodapes = [
        _img_bytes(timbre.get("rodape_img_bin")),
        _img_bytes(timbre.get("rodape_img_2_bin")),
        _img_bytes(timbre.get("rodape_img_3_bin")),
    ]

    box_w = 3.3 * cm
    box_h = 1.35 * cm
    gap = 0.25 * cm
    start_x = w - 2 * cm - (box_w * 3) - (gap * 2)

    for i, img in enumerate(rodapes):
        if img:
            _draw_img_fit(
                c,
                img,
                x=start_x + i * (box_w + gap),
                y=footer_y,
                box_w=box_w,
                box_h=box_h,
            )


# ============================================================
# EXPORT · EXCEL
# ============================================================

def export_pts_excel_service(conn, pts_id: int):
    ensure_pts_schema(conn)

    clinica_id = require_clinica_id()

    pts = fetch_pts_by_id(conn, pts_id, clinica_id)
    if not pts:
        abort(404, "PTS não encontrado.")

    paciente = fetch_paciente_full(conn, pts["paciente_id"], clinica_id)
    participantes = fetch_participantes(conn, pts_id, clinica_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "PTS"

    title_fill = PatternFill("solid", fgColor="1F4E79")
    title_font = Font(color="FFFFFF", bold=True, size=13)
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    section_font = Font(bold=True)

    def add_section(title: str):
        row = ws.max_row + 1
        ws.append([title])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell = ws.cell(row=row, column=1)
        cell.fill = section_fill
        cell.font = section_font

    ws.append(["PLANO TERAPÊUTICO SINGULAR - PTS"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(["PTS", f"#{pts['id']}"])
    ws.append(["Data", pts.get("data_pts", "")])
    ws.append(["Competência", pts.get("competencia", "")])
    ws.append([])

    add_section("Paciente")
    ws.append(["Nome", (paciente or {}).get("nome", "")])
    ws.append(["Prontuário", (paciente or {}).get("prontuario", "")])
    ws.append(["Nascimento", (paciente or {}).get("nascimento", "")])
    ws.append(["Telefone", (paciente or {}).get("telefone", "")])
    ws.append([])

    add_section("Plano terapêutico")
    ws.append(["Objetivo geral", pts.get("objetivo_geral", "")])
    ws.append(["Avaliação", pts.get("avaliacao", "")])
    ws.append(["Plano", pts.get("plano", "")])
    ws.append(["Observações", pts.get("observacoes", "")])
    ws.append([])

    add_section("Participantes")
    ws.append(["Nome", "CBO", "Ocupação/Função"])

    header_row = ws.max_row
    for col in range(1, 4):
        c = ws.cell(row=header_row, column=col)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="EFEFEF")

    if participantes:
        for p in participantes:
            ws.append([
                p.get("nome", ""),
                p.get("cbo", ""),
                p.get("ocupacao") or p.get("funcao", ""),
            ])
    else:
        ws.append(["Sem participantes", "", ""])

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, width in {1: 28, 2: 45, 3: 45}.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    registrar_log(
        conn,
        acao="EXPORTAR_PTS_EXCEL",
        modulo="pts",
        referencia_id=pts_id,
        detalhes=f"Exportação Excel do PTS #{pts_id}",
    )

    conn.commit()

    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"PTS_{pts_id}.xlsx",
    )


# ============================================================
# EXPORT · PDF
# ============================================================

def export_pts_pdf_service(conn, pts_id: int):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.pdfgen import canvas
    except ModuleNotFoundError:
        return (
            "Exportação PDF indisponível: instale reportlab com `pip install reportlab`.",
            501,
        )

    ensure_pts_schema(conn)

    clinica_id = require_clinica_id()

    pts = fetch_pts_by_id(conn, pts_id, clinica_id)
    if not pts:
        abort(404, "PTS não encontrado.")

    paciente = fetch_paciente_full(conn, pts["paciente_id"], clinica_id)
    participantes = fetch_participantes(conn, pts_id, clinica_id)
    timbre = _buscar_timbre_config(conn, clinica_id)

    bio = io.BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    w, h = A4

    margin_x = 2 * cm

    # Espaço extra por causa do cabeçalho do timbre
    top_content_y = h - 4.85 * cm
    bottom_limit = 3.15 * cm
    y = top_content_y

    def draw_page_bg():
        _draw_timbre_page(c, w, h, timbre)

    def new_page():
        nonlocal y
        c.showPage()
        draw_page_bg()
        y = top_content_y

    def ensure_space(space=44):
        if y < bottom_limit + space:
            new_page()

    def line(text, size=10, bold=False, gap=14):
        nonlocal y
        ensure_space(50)
        c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        c.drawString(margin_x, y, str(text or ""))
        y -= gap

    def section(title):
        nonlocal y
        ensure_space(62)
        y -= 5
        c.setFont("Helvetica-Bold", 12)
        c.drawString(margin_x, y, title)
        y -= 6
        c.line(margin_x, y, w - margin_x, y)
        y -= 16

    def wrapped(label, text, size=10):
        nonlocal y
        text = str(text or "").strip()

        section(label)

        if not text:
            line("-", size=size, gap=14)
            return

        max_chars = 95
        partes = []

        while len(text) > max_chars:
            corte = text.rfind(" ", 0, max_chars)
            if corte <= 0:
                corte = max_chars
            partes.append(text[:corte].strip())
            text = text[corte:].strip()

        if text:
            partes.append(text)

        c.setFont("Helvetica", size)

        for parte in partes:
            ensure_space(28)
            c.drawString(margin_x, y, parte)
            y -= 13

        y -= 4

    # Primeira página com timbre
    draw_page_bg()

    # Título do documento
    c.setFont("Helvetica-Bold", 15)
    c.drawString(margin_x, y, "Plano Terapêutico Singular - PTS")

    c.setFont("Helvetica", 10)
    c.drawRightString(w - margin_x, y, f"Data: {pts.get('data_pts', '')}")
    y -= 26

    section("Paciente")

    if paciente:
        line(f"Nome: {paciente.get('nome', '')}")
        line(f"Prontuário: {paciente.get('prontuario', '')}")
        line(f"Nascimento: {paciente.get('nascimento', '')}")
        line(f"Telefone: {paciente.get('telefone', '')}")
    else:
        line("Paciente não encontrado.")

    section("Participantes")

    if participantes:
        for p in participantes:
            nome = p.get("nome", "")
            cbo = p.get("cbo", "")
            ocupacao = p.get("ocupacao") or p.get("funcao", "")

            extra = []
            if ocupacao:
                extra.append(ocupacao)
            if cbo:
                extra.append(f"CBO {cbo}")

            line(f"- {nome}" + (f" · {' · '.join(extra)}" if extra else ""))
    else:
        line("- Sem participantes")

    wrapped("Objetivo geral", pts.get("objetivo_geral", ""))
    wrapped("Avaliação", pts.get("avaliacao", ""))
    wrapped("Plano", pts.get("plano", ""))
    wrapped("Observações", pts.get("observacoes", ""))

    c.showPage()
    c.save()

    bio.seek(0)

    registrar_log(
        conn,
        acao="EXPORTAR_PTS_PDF",
        modulo="pts",
        referencia_id=pts_id,
        detalhes=f"Exportação PDF do PTS #{pts_id} com timbre da clínica {clinica_id}",
    )

    conn.commit()

    return send_file(
        bio,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"PTS_{pts_id}.pdf",
    )