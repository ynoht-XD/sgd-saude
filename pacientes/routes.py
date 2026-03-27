# -*- coding: utf-8 -*-
from __future__ import annotations

import re
import io
import csv
import json
import sqlite3
from datetime import datetime, date
from typing import Any, Dict, List, Tuple, Optional

from flask import (
    render_template, redirect, request, url_for, send_file, jsonify
)

from . import pacientes_bp
from db import conectar_db


# =============================================================================
# SECTION 0 · CONSTANTES / PADRÕES
# =============================================================================

_PT_WEEKDAYS = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]

DEFAULT_TAGS = [
    ("diabetico",  "Diabético"),
    ("bpc",        "BPC"),
    ("cardiopata", "Cardiopata"),
    ("obeso",      "Obeso"),
    ("hipertenso", "Hipertenso"),
    ("cadeirante", "Cadeirante"),
    ("surdo",      "Surdo"),
    ("cego",       "Cego"),
]


# =============================================================================
# SECTION 1 · HELPERS DE SCHEMA (IDEMPOTENTE)
# =============================================================================

def _get_table_columns(table: str) -> set:
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    cols = {row[1] for row in cur.fetchall()}
    conn.close()
    return cols


def _ensure_column(
    conn: sqlite3.Connection,
    table: str,
    col: str,
    ddl_type: str,
    default_sql: str | None = None
) -> None:
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    cols = {r[1] for r in cur.fetchall()}
    if col in cols:
        return

    sql = f"ALTER TABLE {table} ADD COLUMN {col} {ddl_type}"
    if default_sql is not None:
        sql += f" DEFAULT {default_sql}"
    conn.execute(sql)


def ensure_pacientes_schema() -> None:
    with conectar_db() as conn:
        cur = conn.cursor()

        cur.execute("""
            CREATE TABLE IF NOT EXISTS pacientes (
              id INTEGER PRIMARY KEY AUTOINCREMENT
            )
        """)

        _ensure_column(conn, "pacientes", "prontuario", "TEXT", "''")
        _ensure_column(conn, "pacientes", "nome", "TEXT", "''")
        _ensure_column(conn, "pacientes", "nascimento", "TEXT", "''")
        _ensure_column(conn, "pacientes", "idade", "INTEGER", "NULL")
        _ensure_column(conn, "pacientes", "sexo", "TEXT", "''")
        _ensure_column(conn, "pacientes", "status", "TEXT", "''")
        _ensure_column(conn, "pacientes", "mod", "TEXT", "''")
        _ensure_column(conn, "pacientes", "cid", "TEXT", "''")
        _ensure_column(conn, "pacientes", "cid2", "TEXT", "''")

        _ensure_column(conn, "pacientes", "rua", "TEXT", "''")
        _ensure_column(conn, "pacientes", "logradouro", "TEXT", "''")
        _ensure_column(conn, "pacientes", "numero", "TEXT", "''")
        _ensure_column(conn, "pacientes", "numero_casa", "TEXT", "''")
        _ensure_column(conn, "pacientes", "bairro", "TEXT", "''")
        _ensure_column(conn, "pacientes", "cep", "TEXT", "''")
        _ensure_column(conn, "pacientes", "cidade", "TEXT", "''")
        _ensure_column(conn, "pacientes", "municipio", "TEXT", "''")
        _ensure_column(conn, "pacientes", "uf", "TEXT", "''")

        _ensure_column(conn, "pacientes", "cpf", "TEXT", "''")
        _ensure_column(conn, "pacientes", "cns", "TEXT", "''")
        _ensure_column(conn, "pacientes", "telefone", "TEXT", "''")
        _ensure_column(conn, "pacientes", "telefone1", "TEXT", "''")
        _ensure_column(conn, "pacientes", "nome_mae", "TEXT", "''")
        _ensure_column(conn, "pacientes", "mae", "TEXT", "''")
        _ensure_column(conn, "pacientes", "nome_pai", "TEXT", "''")
        _ensure_column(conn, "pacientes", "pai", "TEXT", "''")

        _ensure_column(conn, "pacientes", "end_prontuario", "TEXT", "''")
        _ensure_column(conn, "pacientes", "alergias", "TEXT", "''")
        _ensure_column(conn, "pacientes", "aviso", "TEXT", "''")
        _ensure_column(conn, "pacientes", "comorbidades_json", "TEXT", "'[]'")

        _ensure_column(conn, "pacientes", "terapeuta", "TEXT", "''")
        _ensure_column(conn, "pacientes", "cbo", "TEXT", "''")
        _ensure_column(conn, "pacientes", "cbo_nome", "TEXT", "''")

        conn.commit()


# =============================================================================
# SECTION 2 · HELPERS (NORMALIZAÇÃO / DATAS)
# =============================================================================

_UPPER_FIELDS = {
    "nome", "mae", "pai", "responsavel",
    "logradouro", "rua", "bairro", "municipio", "cidade", "complemento",
    "estado_civil",
    "orgao_rg", "orgao_rg_responsavel",
    "status", "mod", "cid", "cid2", "raca",
    "codigo_logradouro",
    "terapeuta", "cbo", "cbo_nome",
}

def _to_upper(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip().upper()

def _upperize_payload(dados: dict) -> dict:
    out = {}
    for k, v in dados.items():
        out[k] = _to_upper(v) if k in _UPPER_FIELDS else v
    return out

def _calc_idade(nasc_str: Any) -> Optional[int]:
    if not nasc_str:
        return None

    nasc_str = str(nasc_str).strip()
    fmts = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d")
    dt = None

    for f in fmts:
        try:
            dt = datetime.strptime(nasc_str, f).date()
            break
        except Exception:
            continue

    if not dt:
        s = "".join(ch for ch in nasc_str if ch.isdigit())
        try:
            if len(s) == 8:
                if int(s[:4]) > 1900:
                    dt = date(int(s[:4]), int(s[4:6]), int(s[6:8]))
                else:
                    dt = date(int(s[4:8]), int(s[2:4]), int(s[0:2]))
        except Exception:
            return None

    if not dt:
        return None

    today = date.today()
    anos = today.year - dt.year - ((today.month, today.day) < (dt.month, dt.day))
    return max(0, anos)

def _parse_dt_flex(s: Any) -> Optional[datetime]:
    if not s:
        return None

    s = str(s).strip()
    if not s:
        return None

    s2 = s.replace("Z", "").strip()

    fmts = (
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M",

        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
    )

    for fmt in fmts:
        try:
            return datetime.strptime(s2, fmt)
        except Exception:
            pass

    try:
        return datetime.fromisoformat(s2)
    except Exception:
        pass

    m = re.search(r"(\d{4})-(\d{2})-(\d{2}).*?(\d{2}):(\d{2})", s2)
    if m:
        try:
            y, mo, d, hh, mm = map(int, m.groups())
            return datetime(y, mo, d, hh, mm, 0)
        except Exception:
            return None

    m2 = re.search(r"(\d{2})/(\d{2})/(\d{4}).*?(\d{2}):(\d{2})", s2)
    if m2:
        try:
            d, mo, y, hh, mm = map(int, m2.groups())
            return datetime(y, mo, d, hh, mm, 0)
        except Exception:
            return None

    return None

def _enriquecer_agendamento_row(a: dict) -> dict:
    dt_ini = _parse_dt_flex(a.get("inicio"))
    dt_fim = _parse_dt_flex(a.get("fim"))

    if dt_ini:
        a["dia_semana"] = _PT_WEEKDAYS[dt_ini.weekday()]
        a["hora_ini"]   = dt_ini.strftime("%H:%M")
        a["data_br"]    = dt_ini.strftime("%d/%m/%Y")
    else:
        a["dia_semana"] = ""
        a["hora_ini"]   = ""
        a["data_br"]    = ""

    a["hora_fim"] = dt_fim.strftime("%H:%M") if dt_fim else ""
    return a

def _json_list(v: Any) -> List[str]:
    if v is None:
        return []
    if isinstance(v, list):
        return [str(x).strip() for x in v if str(x).strip()]
    s = str(v).strip()
    if not s:
        return []
    try:
        j = json.loads(s)
        if isinstance(j, list):
            return [str(x).strip() for x in j if str(x).strip()]
    except Exception:
        pass
    return [p.strip() for p in s.split(",") if p.strip()]


# =============================================================================
# SECTION 3 · HELPERS (PROFISSIONAIS → CBO) + AGENDAMENTOS
# =============================================================================

_SPLIT_PROF_RE = re.compile(r"\s*(?:,|;|/|\||\+|&|\be\b)\s*", re.IGNORECASE)

def _split_profissionais(raw: str) -> list[str]:
    if not raw:
        return []
    s = str(raw).strip()
    if not s:
        return []
    partes = [p.strip() for p in _SPLIT_PROF_RE.split(s) if p and p.strip()]
    seen = set()
    out = []
    for p in partes:
        k = p.upper()
        if k in seen:
            continue
        seen.add(k)
        out.append(p)
    return out

def _map_cbo_por_profissionais(nomes: list[str]) -> dict[str, str]:
    if not nomes:
        return {}

    cols = _get_table_columns("usuarios")
    if "nome" not in cols or "cbo" not in cols:
        return {}

    ph = ",".join(["?"] * len(nomes))
    sql = f"""
        SELECT nome, cbo
          FROM usuarios
         WHERE TRIM(COALESCE(nome,'')) <> ''
           AND nome COLLATE NOCASE IN ({ph})
    """

    conn = conectar_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(sql, nomes)
    rows = cur.fetchall()
    conn.close()

    m = {}
    for r in rows:
        nm = (r["nome"] or "").strip()
        cbo = (r["cbo"] or "").strip()
        if nm:
            m[nm.upper()] = cbo
    return m

def _enriquecer_com_prof_cbo(a: dict, cbo_map: dict[str, str]) -> dict:
    prof_raw = a.get("profissional") or ""
    profs = _split_profissionais(prof_raw)

    cbos: list[str] = []
    seen = set()
    for p in profs:
        cbo = (cbo_map.get(p.upper(), "") or "").strip()
        if cbo and cbo not in seen:
            seen.add(cbo)
            cbos.append(cbo)

    a["profissionais_lista"] = profs
    a["cbo_lista"] = cbos
    a["cbo_str"] = ", ".join(cbos) if cbos else ""
    return a

def _get_primeiro_agendamento_por_paciente() -> dict:
    cols_ag = _get_table_columns("agendamentos")
    if "paciente" not in cols_ag or "inicio" not in cols_ag:
        return {}

    has_fim  = "fim" in cols_ag
    has_prof = "profissional" in cols_ag

    sel = ["paciente", "inicio"]
    sel.append("fim" if has_fim else "NULL AS fim")
    sel.append("profissional" if has_prof else "'' AS profissional")

    conn = conectar_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(f"""
        SELECT {", ".join(sel)}
          FROM agendamentos
         WHERE TRIM(COALESCE(paciente,'')) <> ''
    """)
    rows = cur.fetchall()
    conn.close()

    now = datetime.now()

    por_paciente: dict[str, dict] = {}
    todos_profissionais: list[str] = []

    for r in rows:
        pac = (r["paciente"] or "").strip()
        if not pac:
            continue
        k = pac.upper()

        por_paciente.setdefault(k, {
            "slots_by_key": {},   # key -> {"slot": dict_enriquecido, "dt": datetime}
            "terapeutas": set(),
        })

        profs = _split_profissionais(r["profissional"] or "")
        for p in profs:
            por_paciente[k]["terapeutas"].add(p)
            todos_profissionais.append(p)

        dt_ini = _parse_dt_flex(r["inicio"])
        if not dt_ini:
            continue

        slot = _enriquecer_agendamento_row({"inicio": r["inicio"], "fim": r["fim"]})

        dia = (slot.get("dia_semana") or "").strip()
        hi  = (slot.get("hora_ini") or "").strip()
        hf  = (slot.get("hora_fim") or "").strip()
        if not dia or not hi:
            continue

        key_slot = f"{dia}|{hi}|{hf}"

        entry = por_paciente[k]["slots_by_key"].get(key_slot)
        if entry is None:
            por_paciente[k]["slots_by_key"][key_slot] = {"slot": slot, "dt": dt_ini}
        else:
            dt_old = entry["dt"]
            old_future = dt_old >= now
            new_future = dt_ini >= now

            if old_future and new_future:
                if dt_ini < dt_old:
                    por_paciente[k]["slots_by_key"][key_slot] = {"slot": slot, "dt": dt_ini}
            elif (not old_future) and new_future:
                por_paciente[k]["slots_by_key"][key_slot] = {"slot": slot, "dt": dt_ini}
            elif (not old_future) and (not new_future):
                if dt_ini < dt_old:
                    por_paciente[k]["slots_by_key"][key_slot] = {"slot": slot, "dt": dt_ini}

    uniq_profs, seen = [], set()
    for p in todos_profissionais:
        up = p.upper()
        if up not in seen:
            seen.add(up)
            uniq_profs.append(p)

    cbo_map = _map_cbo_por_profissionais(uniq_profs)

    order_dia = {d: i for i, d in enumerate(_PT_WEEKDAYS)}

    resultado: dict[str, dict] = {}
    for pac, info in por_paciente.items():
        terapeutas = sorted(info["terapeutas"])

        cbos: list[str] = []
        seen_cbo = set()
        for t in terapeutas:
            cbo = (cbo_map.get(t.upper(), "") or "").strip()
            if cbo and cbo not in seen_cbo:
                seen_cbo.add(cbo)
                cbos.append(cbo)

        slots = [v["slot"] for v in info["slots_by_key"].values()]
        slots.sort(key=lambda s: (
            order_dia.get(s.get("dia_semana", ""), 99),
            s.get("hora_ini", "99:99")
        ))

        prox_slot = None
        prox_dt = None
        for entry in info["slots_by_key"].values():
            dt_ini = entry["dt"]
            if dt_ini >= now and (prox_dt is None or dt_ini < prox_dt):
                prox_dt = dt_ini
                prox_slot = entry["slot"]

        agenda_partes = []
        for s in slots:
            dia = s.get("dia_semana", "")
            hi  = s.get("hora_ini", "")
            hf  = s.get("hora_fim", "")
            if hf:
                agenda_partes.append(f"{dia[:3].upper()} {hi}–{hf}")
            else:
                agenda_partes.append(f"{dia[:3].upper()} {hi}")

        prox = prox_slot or {}

        resultado[pac] = {
            "dia_semana": prox.get("dia_semana", ""),
            "hora_ini": prox.get("hora_ini", ""),
            "hora_fim": prox.get("hora_fim", ""),

            "agenda_lista": slots,
            "agenda_str": "; ".join(agenda_partes),

            "terapeutas": terapeutas,
            "cbo_lista": cbos,
            "terapeuta_str": " / ".join(terapeutas),
            "cbo_str": ", ".join(cbos),
        }

    return resultado


def _fetch_agendamentos_por_paciente(nome_paciente: str) -> dict:
    if not nome_paciente:
        return {"agds_upcoming": [], "agds_all": [], "series_resumo": [],
                "total_agds": 0, "total_upcoming": 0}

    cols_ag = _get_table_columns("agendamentos")
    if "paciente" not in cols_ag or "inicio" not in cols_ag:
        return {"agds_upcoming": [], "agds_all": [], "series_resumo": [],
                "total_agds": 0, "total_upcoming": 0}

    sel = ["id", "paciente", "inicio"]
    sel.append("fim" if "fim" in cols_ag else "NULL AS fim")
    sel.append("profissional" if "profissional" in cols_ag else "'' AS profissional")
    sel.append("observacao" if "observacao" in cols_ag else "'' AS observacao")
    sel.append("status" if "status" in cols_ag else "'' AS status")
    sel.append("recorrente" if "recorrente" in cols_ag else "0 AS recorrente")
    sel.append("serie_uid" if "serie_uid" in cols_ag else "'' AS serie_uid")
    sel.append("profissional_cpf" if "profissional_cpf" in cols_ag else "'' AS profissional_cpf")
    sel.append("dow_dom" if "dow_dom" in cols_ag else "'' AS dow_dom")

    conn = conectar_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute(f"""
        SELECT {", ".join(sel)}
          FROM agendamentos
         WHERE paciente = :nome COLLATE NOCASE
         ORDER BY datetime(inicio) DESC
    """, {"nome": nome_paciente})
    agds_all = [_enriquecer_agendamento_row(dict(r)) for r in cur.fetchall()]

    cur.execute(f"""
        SELECT {", ".join(sel)}
          FROM agendamentos
         WHERE paciente = :nome COLLATE NOCASE
           AND datetime(inicio) >= datetime('now','localtime')
         ORDER BY datetime(inicio) ASC
    """, {"nome": nome_paciente})
    agds_upcoming = [_enriquecer_agendamento_row(dict(r)) for r in cur.fetchall()]

    if "serie_uid" in cols_ag:
        cur.execute("""
            SELECT COALESCE(serie_uid, '') AS serie_uid,
                   COALESCE(profissional,'') AS profissional,
                   COUNT(*) AS total_sessoes,
                   SUM(CASE WHEN datetime(inicio) >= datetime('now','localtime') THEN 1 ELSE 0 END) AS sessoes_futuras
              FROM agendamentos
             WHERE paciente = :nome COLLATE NOCASE
             GROUP BY COALESCE(serie_uid, ''), COALESCE(profissional,'')
             ORDER BY MIN(datetime(inicio)) ASC
        """, {"nome": nome_paciente})
        series_resumo = [dict(r) for r in cur.fetchall()]
    else:
        series_resumo = []

    conn.close()

    profs_all: list[str] = []
    for a in (agds_upcoming + agds_all):
        profs_all.extend(_split_profissionais(a.get("profissional") or ""))

    uniq = []
    seen = set()
    for p in profs_all:
        up = p.upper()
        if up in seen:
            continue
        seen.add(up)
        uniq.append(p)

    cbo_map = _map_cbo_por_profissionais(uniq)

    for a in agds_upcoming:
        _enriquecer_com_prof_cbo(a, cbo_map)
    for a in agds_all:
        _enriquecer_com_prof_cbo(a, cbo_map)

    prox_por_serie = {}
    for a in agds_upcoming:
        key = (a.get("serie_uid") or "", a.get("profissional") or "")
        if key not in prox_por_serie:
            prox_por_serie[key] = a
    for a in agds_all:
        key = (a.get("serie_uid") or "", a.get("profissional") or "")
        if key not in prox_por_serie:
            prox_por_serie[key] = a

    for s in series_resumo:
        key = (s.get("serie_uid") or "", s.get("profissional") or "")
        amostra = prox_por_serie.get(key)
        s["amostra_dia_semana"] = amostra.get("dia_semana") if amostra else None
        if amostra:
            s["amostra_hora"] = amostra["hora_ini"] + (f"–{amostra['hora_fim']}" if amostra.get("hora_fim") else "")
        else:
            s["amostra_hora"] = None

    return {
        "agds_upcoming": agds_upcoming,
        "agds_all": agds_all,
        "series_resumo": series_resumo,
        "total_agds": len(agds_all),
        "total_upcoming": len(agds_upcoming),
    }


# =============================================================================
# SECTION 4 · HELPERS (FILTROS / FETCH PACIENTES)
# =============================================================================

def _where_and_params(args, cols: set) -> Tuple[str, Dict[str, Any]]:
    where = []
    params: Dict[str, Any] = {}

    pront = (args.get("prontuario") or "").strip()
    if pront and "prontuario" in cols:
        where.append("prontuario = :pront")
        params["pront"] = pront

    nome = (args.get("nome") or "").strip()
    if nome and "nome" in cols:
        where.append("nome LIKE :nome COLLATE NOCASE")
        params["nome"] = f"%{nome}%"

    sexo = (args.get("sexo") or "").strip().upper()
    if sexo in ("M", "F") and "sexo" in cols:
        where.append("sexo = :sexo")
        params["sexo"] = sexo

    status = (args.get("status") or "").strip()
    if status and "status" in cols:
        where.append("status = :status")
        params["status"] = status

    mod = (args.get("mod") or "").strip()
    if mod and "mod" in cols:
        where.append("mod LIKE :mod COLLATE NOCASE")
        params["mod"] = f"%{mod}%"

    cid = (args.get("cid") or "").strip()
    if cid and ("cid" in cols or "cid2" in cols):
        if "cid" in cols and "cid2" in cols:
            where.append("(cid LIKE :cid OR cid2 LIKE :cid) COLLATE NOCASE")
        elif "cid" in cols:
            where.append("cid LIKE :cid COLLATE NOCASE")
        else:
            where.append("cid2 LIKE :cid COLLATE NOCASE")
        params["cid"] = f"%{cid}%"

    cidade_val = (args.get("cidade") or "").strip()
    if cidade_val:
        if "cidade" in cols:
            where.append("cidade LIKE :cidade COLLATE NOCASE")
            params["cidade"] = f"%{cidade_val}%"
        elif "municipio" in cols:
            where.append("municipio LIKE :cidade COLLATE NOCASE")
            params["cidade"] = f"%{cidade_val}%"

    bairro = (args.get("bairro") or "").strip()
    if bairro and "bairro" in cols:
        where.append("bairro LIKE :bairro COLLATE NOCASE")
        params["bairro"] = f"%{bairro}%"

    rua = (args.get("rua") or "").strip()
    if rua:
        parts = []
        if "rua" in cols:
            parts.append("rua LIKE :rua COLLATE NOCASE")
        if "logradouro" in cols:
            parts.append("logradouro LIKE :rua COLLATE NOCASE")
        if parts:
            where.append("(" + " OR ".join(parts) + ")")
            params["rua"] = f"%{rua}%"

    mes_nasc = (args.get("mes_nasc") or "").strip()
    if mes_nasc.isdigit() and "nascimento" in cols:
        mes2 = mes_nasc.zfill(2)
        where.append("""
            (
              SUBSTR(nascimento, 6, 2) = :mes2
              OR SUBSTR(nascimento, 4, 2) = :mes2
              OR SUBSTR(nascimento, 5, 2) = :mes2
            )
        """)
        params["mes2"] = mes2

    clause = " WHERE " + " AND ".join(where) if where else ""
    return clause, params


def _fetch_pacientes_list(args=None):
    ensure_pacientes_schema()
    cols = _get_table_columns("pacientes")

    base_cols = [
        "id", "prontuario", "nome", "nascimento", "idade", "sexo", "status", "mod", "cid", "cid2",
        "cpf", "cns", "telefone", "telefone1",
        "nome_mae", "mae", "nome_pai", "pai",
        "rua", "logradouro", "numero", "numero_casa", "bairro", "cep", "cidade", "municipio", "uf",
        "end_prontuario", "alergias", "aviso", "comorbidades_json",
    ]
    select_cols = [c for c in base_cols if c in cols]
    if not select_cols:
        select_cols = ["*"]

    if args is None:
        clause, params = "", {}
    else:
        clause, params = _where_and_params(args, cols)

    conn = conectar_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(
        f"SELECT {', '.join(select_cols)} FROM pacientes {clause} ORDER BY id DESC",
        params
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()

    for r in rows:
        if not (r.get("telefone") or "").strip():
            r["telefone"] = (r.get("telefone1") or "").strip()
        if not (r.get("nome_mae") or "").strip():
            r["nome_mae"] = (r.get("mae") or "").strip()
        if not (r.get("nome_pai") or "").strip():
            r["nome_pai"] = (r.get("pai") or "").strip()

    if args:
        idade_min = args.get("idade_min", type=int)
        idade_max = args.get("idade_max", type=int)
        if idade_min is not None or idade_max is not None:
            filtradas = []
            for r in rows:
                idade = r.get("idade")
                if idade is None:
                    idade = _calc_idade(r.get("nascimento"))
                if idade is None:
                    continue
                if idade_min is not None and idade < idade_min:
                    continue
                if idade_max is not None and idade > idade_max:
                    continue
                r["idade"] = idade
                filtradas.append(r)
            rows = filtradas

    mapa_ag = _get_primeiro_agendamento_por_paciente()

    for r in rows:
        key = (r.get("nome") or "").strip().upper()
        info = mapa_ag.get(key)

        r["ag_dia"] = (info.get("dia_semana") if info else "") or ""
        r["ag_hora_ini"] = (info.get("hora_ini") if info else "") or ""
        r["ag_hora_fim"] = (info.get("hora_fim") if info else "") or ""
        r["ag_resumo"] = (info.get("agenda_str") if info else "") or ""

        r["terapeuta"] = (info.get("terapeuta_str") if info else "") or ""
        r["cbo"]       = (info.get("cbo_str") if info else "") or ""
        r["cbo_nome"]  = ""

    # >>>>>>>>>>>> AQUI: filtros finais usando os nomes corretos do teu HTML <<<<<<<<<<<<
    if args:
        # teu HTML manda: name="dia_semana"
        dia_semana = (args.get("dia_semana") or "").strip().lower()
        terapeuta_q = (args.get("terapeuta") or "").strip().lower()
        cbo_q = (args.get("cbo") or "").strip().lower()

        if dia_semana:
            rows = [r for r in rows if (r.get("ag_dia") or "").strip().lower() == dia_semana]

        if terapeuta_q:
            rows = [r for r in rows if terapeuta_q in (r.get("terapeuta") or "").lower()]

        if cbo_q:
            rows = [r for r in rows if cbo_q in (r.get("cbo") or "").lower()]

    return rows


def _headers_padrao():
    return ["ID", "Prontuário", "Nome", "Nascimento", "Sexo", "Status", "Modalidade", "CID"]


# =============================================================================
# SECTION 5 · ROTAS PRINCIPAIS (TELAS)
# =============================================================================

@pacientes_bp.route("/")
def listar_pacientes():
    rows = _fetch_pacientes_list(request.args)
    return render_template("pacientes.html", pacientes=rows)

@pacientes_bp.route("/pacientes")
def listar_pacientes_compat():
    return redirect(url_for("pacientes.listar_pacientes"), code=302)


@pacientes_bp.route("/visualizar/<int:id>")
def visualizar_paciente(id):
    ensure_pacientes_schema()

    conn = conectar_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM pacientes WHERE id = ? LIMIT 1", (id,))
    row = cur.fetchone()
    conn.close()

    if not row:
        return "Paciente não encontrado", 404

    paciente = dict(row)
    agds = _fetch_agendamentos_por_paciente(paciente.get("nome", ""))

    return render_template(
        "visualizar_paciente.html",
        paciente=paciente,
        agds_upcoming=agds["agds_upcoming"],
        agds_all=agds["agds_all"],
        series_resumo=agds["series_resumo"],
        total_agds=agds["total_agds"],
        total_upcoming=agds["total_upcoming"],
    )


@pacientes_bp.route("/editar/<int:id>")
def editar_paciente(id):
    ensure_pacientes_schema()

    conn = conectar_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM pacientes WHERE id = ?", (id,))
    paciente = cur.fetchone()
    conn.close()

    if not paciente:
        return "Paciente não encontrado", 404

    return render_template("editar_paciente.html", paciente=paciente)


@pacientes_bp.route("/atualizar/<int:id>", methods=["POST"])
def atualizar_paciente(id: int):
    ensure_pacientes_schema()

    dados_raw = request.form.to_dict(flat=True)
    dados = _upperize_payload(dados_raw)

    cols = _get_table_columns("pacientes")
    pairs: list[tuple[str, Any]] = []

    def add_if_exists(col: str, val: Any):
        if col in cols:
            # normaliza None -> "" pra não quebrar visual
            pairs.append((col, "" if val is None else val))

    # --------- principais ----------
    add_if_exists("status", dados.get("status"))
    add_if_exists("mod", dados.get("mod"))
    add_if_exists("nome", dados.get("nome"))
    add_if_exists("nascimento", dados.get("nascimento"))
    add_if_exists("sexo", (dados.get("sexo") or "").strip().upper())
    add_if_exists("cid", dados.get("cid"))
    add_if_exists("cid2", dados.get("cid2"))
    add_if_exists("admissao", dados.get("admissao"))
    add_if_exists("raca", dados.get("raca"))

    # idade: recalcula (não confia em readonly)
    nasc = (dados.get("nascimento") or "").strip()
    idade_calc = _calc_idade(nasc) if nasc else None
    if "idade" in cols:
        pairs.append(("idade", idade_calc))

    # --------- endereço ----------
    add_if_exists("logradouro", dados.get("logradouro"))
    add_if_exists("bairro", dados.get("bairro"))
    add_if_exists("numero_casa", dados.get("numero_casa"))
    add_if_exists("complemento", dados.get("complemento"))
    add_if_exists("cep", dados.get("cep"))
    add_if_exists("municipio", dados.get("municipio"))
    add_if_exists("codigo_logradouro", dados.get("codigo_logradouro"))

    # compat (se teus templates usam rua/numero/cidade em algum lugar)
    # mantém sincronizado quando existir
    if "rua" in cols and not (dados.get("rua") or "").strip():
        pairs.append(("rua", dados.get("logradouro") or ""))
    if "numero" in cols and not (dados.get("numero") or "").strip():
        pairs.append(("numero", dados.get("numero_casa") or ""))
    if "cidade" in cols and not (dados.get("cidade") or "").strip():
        pairs.append(("cidade", dados.get("municipio") or ""))

    # --------- documentos ----------
    add_if_exists("cpf", dados.get("cpf"))
    add_if_exists("cns", dados.get("cns"))
    add_if_exists("estado_civil", dados.get("estado_civil"))
    add_if_exists("rg", dados.get("rg"))
    add_if_exists("orgao_rg", dados.get("orgao_rg"))
    add_if_exists("nis", dados.get("nis"))

    # --------- contatos ----------
    add_if_exists("telefone1", dados.get("telefone1"))
    add_if_exists("telefone2", dados.get("telefone2"))
    add_if_exists("telefone3", dados.get("telefone3"))
    add_if_exists("email", dados.get("email"))

    # compat telefone (se algum lugar usa "telefone")
    if "telefone" in cols and not (dados.get("telefone") or "").strip():
        pairs.append(("telefone", dados.get("telefone1") or ""))

    # --------- familiares ----------
    add_if_exists("mae", dados.get("mae"))
    add_if_exists("cpf_mae", dados.get("cpf_mae"))
    add_if_exists("rg_mae", dados.get("rg_mae"))
    add_if_exists("rg_ssp_mae", dados.get("rg_ssp_mae"))
    add_if_exists("nis_mae", dados.get("nis_mae"))

    add_if_exists("pai", dados.get("pai"))
    add_if_exists("cpf_pai", dados.get("cpf_pai"))
    add_if_exists("rg_pai", dados.get("rg_pai"))
    add_if_exists("rg_ssp_pai", dados.get("rg_ssp_pai"))

    # --------- responsável ----------
    add_if_exists("responsavel", dados.get("responsavel"))
    add_if_exists("cpf_responsavel", dados.get("cpf_responsavel"))
    add_if_exists("rg_responsavel", dados.get("rg_responsavel"))
    add_if_exists("orgao_rg_responsavel", dados.get("orgao_rg_responsavel"))

    # se nada pra atualizar, volta
    if not pairs:
        return redirect(url_for("pacientes.visualizar_paciente", id=id))

    # remove duplicados (mantém o último valor se repetir)
    dedup = {}
    for k, v in pairs:
        dedup[k] = v
    pairs = list(dedup.items())

    set_sql = ", ".join([f"{c} = ?" for c, _ in pairs])
    vals = [v for _, v in pairs] + [id]

    try:
        with conectar_db() as conn:
            # debug: confirma qual banco está sendo usado
            # print("DB =>", conn.execute("PRAGMA database_list").fetchall())

            cur = conn.cursor()
            cur.execute(f"UPDATE pacientes SET {set_sql} WHERE id = ?", vals)
            conn.commit()

            # se rowcount = 0, ou id não existe, ou valores iguais
            # (ainda assim é útil saber)
            # print("UPDATE rowcount:", cur.rowcount)

        return redirect(url_for("pacientes.visualizar_paciente", id=id))

    except Exception as e:
        return f"Erro ao atualizar paciente: {e}", 500


# =============================================================================
# SECTION 6 · API (AGENDAMENTOS DO PACIENTE)
# =============================================================================

@pacientes_bp.route("/api/paciente/<int:id>/agendamentos")
def api_agendamentos_paciente(id):
    conn = conectar_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT nome FROM pacientes WHERE id = ? LIMIT 1", (id,))
    r = cur.fetchone()
    conn.close()

    if not r:
        return jsonify({"erro": "Paciente não encontrado"}), 404

    data = _fetch_agendamentos_por_paciente(r["nome"])
    return jsonify(data)


# =============================================================================
# SECTION 7 · API (AUTOSAVE DO CARD)
# =============================================================================

@pacientes_bp.route("/api/autosave", methods=["POST"])
def api_autosave():
    ensure_pacientes_schema()
    payload = request.get_json(silent=True) or {}

    pid = payload.get("id")
    field = (payload.get("field") or "").strip()
    value = payload.get("value")

    if not pid:
        return jsonify({"error": "id obrigatório"}), 400

    allowed = {
        "end_prontuario": "end_prontuario",
        "alergias": "alergias",
        "aviso": "aviso",
        "tags": "comorbidades_json",
    }
    if field not in allowed:
        return jsonify({"error": f"field inválido: {field}"}), 400

    col = allowed[field]
    if field == "tags":
        arr = _json_list(value)
        value_to_save = json.dumps(arr, ensure_ascii=False)
    else:
        value_to_save = "" if value is None else str(value)

    try:
        with conectar_db() as conn:
            conn.execute(f"UPDATE pacientes SET {col} = ? WHERE id = ?", (value_to_save, pid))
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# =============================================================================
# SECTION 8 · APIs DE SUGESTÕES (TYPEAHEAD/DATALIST)
# =============================================================================



@pacientes_bp.route("/api/sugestoes/prontuarios")
def api_sugestoes_prontuarios():
    """Sugestões de prontuário (>= 3 chars)."""
    ensure_pacientes_schema()

    q = (request.args.get("q") or "").strip()
    if len(q) < 3:
        return jsonify([])

    cols = _get_table_columns("pacientes")
    if "prontuario" not in cols:
        return jsonify([])

    conn = conectar_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT prontuario
          FROM pacientes
         WHERE TRIM(COALESCE(prontuario,'')) <> ''
           AND prontuario LIKE :q
         ORDER BY prontuario
         LIMIT 20
    """, {"q": f"%{q}%"})
    rows = cur.fetchall()
    conn.close()

    return jsonify([r["prontuario"] for r in rows if r["prontuario"]])


@pacientes_bp.route("/api/sugestoes/nomes")
def api_sugestoes_nomes():
    """Sugestões de nome (>= 3 chars). Retorna JSON rico (nome/cpf/idade/prontuario)."""
    ensure_pacientes_schema()

    q = (request.args.get("q") or "").strip()
    if len(q) < 3:
        return jsonify([])

    cols = _get_table_columns("pacientes")
    have_cpf = "cpf" in cols
    have_pront = "prontuario" in cols
    have_nasc = "nascimento" in cols

    sel_parts = ["nome"]
    sel_parts.append("nascimento" if have_nasc else "NULL AS nascimento")
    sel_parts.append("cpf" if have_cpf else "NULL AS cpf")
    sel_parts.append("prontuario" if have_pront else "NULL AS prontuario")

    sql = f"""
        SELECT DISTINCT {", ".join(sel_parts)}
          FROM pacientes
         WHERE nome LIKE :q COLLATE NOCASE
         ORDER BY nome
         LIMIT 20
    """

    conn = conectar_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(sql, {"q": f"%{q}%"})
    rows = cur.fetchall()
    conn.close()

    out = []
    for r in rows:
        nasc = r["nascimento"] if "nascimento" in r.keys() else None
        out.append({
            "nome": (r["nome"] or "").strip(),
            "cpf": (r["cpf"] or "") if "cpf" in r.keys() else "",
            "idade": _calc_idade(nasc) if nasc else None,
            "prontuario": (r["prontuario"] or "") if "prontuario" in r.keys() else "",
        })
    return jsonify(out)


@pacientes_bp.route("/api/sugestoes/terapeutas")
def api_sugestoes_terapeutas():
    """
    Sugestões de terapeutas vindos de agendamentos.profissional.
    - Dispara com >= 3 chars
    - Quebra nomes por separadores (virgula, /, ;, ' e ')
    """
    q = (request.args.get("q") or "").strip()
    if len(q) < 3:
        return jsonify([])

    cols_ag = _get_table_columns("agendamentos")
    if "profissional" not in cols_ag:
        return jsonify([])

    conn = conectar_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT profissional
          FROM agendamentos
         WHERE TRIM(COALESCE(profissional,'')) <> ''
           AND profissional LIKE :q COLLATE NOCASE
         ORDER BY profissional
         LIMIT 50
    """, {"q": f"%{q}%"})
    rows = cur.fetchall()
    conn.close()

    nomes = set()
    for r in rows:
        raw = (r["profissional"] or "").strip()
        for p in _split_profissionais(raw):
            if len(p.strip()) >= 3 and q.lower() in p.lower():
                nomes.add(p.strip())

    return jsonify(sorted(nomes)[:20])


@pacientes_bp.route("/api/sugestoes/cids")
def api_sugestoes_cids():
    ensure_pacientes_schema()
    q = (request.args.get("q") or "").strip()

    conn = conectar_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if q:
        cur.execute("""
            SELECT DISTINCT cid
              FROM pacientes
             WHERE cid LIKE :q COLLATE NOCASE
             ORDER BY cid
             LIMIT 20
        """, {"q": f"%{q}%"})
    else:
        cur.execute("""
            SELECT DISTINCT cid
              FROM pacientes
             WHERE cid IS NOT NULL AND TRIM(cid) <> ''
             ORDER BY cid
             LIMIT 20
        """)
    rows = cur.fetchall()
    conn.close()
    return jsonify([r["cid"] for r in rows if r["cid"]])


@pacientes_bp.route("/api/sugestoes/modalidades")
def api_sugestoes_modalidades():
    ensure_pacientes_schema()
    q = (request.args.get("q") or "").strip()

    conn = conectar_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if q:
        cur.execute("""
            SELECT DISTINCT mod
              FROM pacientes
             WHERE mod LIKE :q COLLATE NOCASE
             ORDER BY mod
             LIMIT 20
        """, {"q": f"%{q}%"})
    else:
        cur.execute("""
            SELECT DISTINCT mod
              FROM pacientes
             WHERE mod IS NOT NULL AND TRIM(mod) <> ''
             ORDER BY mod
             LIMIT 20
        """)
    rows = cur.fetchall()
    conn.close()
    return jsonify([r["mod"] for r in rows if r["mod"]])


# =============================================================================
# SECTION 9 · EXPORTS (TURBO: exporta "tudo que existir" + campos derivados)
# =============================================================================

def _export_header_order() -> list[str]:
    """
    Ordem preferida das colunas no arquivo.
    O que não existir no row vai ser ignorado; o resto entra no final.
    """
    return [
        "id",
        "prontuario",
        "nome",
        "nascimento",
        "idade",
        "sexo",
        "cpf",
        "cns",
        "telefone",
        "status",
        "mod",
        "cid",
        "cid2",

        # endereço
        "rua",
        "logradouro",
        "numero",
        "numero_casa",
        "bairro",
        "cep",
        "cidade",
        "municipio",
        "uf",

        # família
        "nome_mae",
        "mae",
        "nome_pai",
        "pai",

        # card / autosave
        "end_prontuario",
        "alergias",
        "aviso",
        "comorbidades_json",

        # derivados do agendamento (enriquecidos em _fetch_pacientes_list)
        "terapeuta",
        "cbo",
        "ag_dia",
        "ag_hora_ini",
        "ag_hora_fim",
        "ag_resumo",
    ]


def _pretty_header(col: str) -> str:
    """
    Converte chave do dict em título amigável.
    """
    mapa = {
        "id": "ID",
        "prontuario": "Prontuário",
        "nome": "Nome",
        "nascimento": "Nascimento",
        "idade": "Idade",
        "sexo": "Sexo",
        "cpf": "CPF",
        "cns": "CNS",
        "telefone": "Telefone",
        "status": "Status",
        "mod": "Modalidade",
        "cid": "CID",
        "cid2": "CID 2",

        "rua": "Rua",
        "logradouro": "Logradouro",
        "numero": "Número",
        "numero_casa": "Número (casa)",
        "bairro": "Bairro",
        "cep": "CEP",
        "cidade": "Cidade",
        "municipio": "Município",
        "uf": "UF",

        "nome_mae": "Nome da mãe",
        "mae": "Mãe",
        "nome_pai": "Nome do pai",
        "pai": "Pai",

        "end_prontuario": "END (Prontuário físico)",
        "alergias": "Alergias",
        "aviso": "Aviso / Situação",
        "comorbidades_json": "Comorbidades (JSON)",

        "terapeuta": "Terapeuta(s)",
        "cbo": "CBO(s)",
        "ag_dia": "Agendamento (Dia)",
        "ag_hora_ini": "Agendamento (Início)",
        "ag_hora_fim": "Agendamento (Fim)",
        "ag_resumo": "Agendamento (Resumo)",
    }
    return mapa.get(col, col.replace("_", " ").strip().title())


def _normalize_cell_value(v: Any) -> str:
    """
    Normaliza valores para Excel/CSV.
    - dict/list vira JSON
    - None vira ""
    """
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        try:
            return json.dumps(v, ensure_ascii=False)
        except Exception:
            return str(v)
    return str(v)


@pacientes_bp.route("/exportar_xls")
def exportar_xls():
    rows = _fetch_pacientes_list(request.args)

    # --- Descobre TODAS as colunas existentes nos rows (dinâmico) ---
    keys_all: set[str] = set()
    for r in rows:
        if isinstance(r, dict):
            keys_all.update(r.keys())

    # ordem preferida + resto
    preferred = _export_header_order()
    cols = [c for c in preferred if c in keys_all]
    resto = sorted([c for c in keys_all if c not in cols])
    cols.extend(resto)

    # headers bonitos
    headers = [_pretty_header(c) for c in cols]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Pacientes"

        # header
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # linhas
        for r in rows:
            ws.append([_normalize_cell_value(r.get(c)) for c in cols])

        # auto width (com limite)
        for idx, col_name in enumerate(cols, start=1):
            letter = get_column_letter(idx)
            max_len = len(headers[idx - 1])
            for cell in ws[letter]:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_len + 2, 60)

        # congela cabeçalho
        ws.freeze_panes = "A2"

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"pacientes_full_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except ImportError:
        bio = io.StringIO()
        writer = csv.writer(bio, delimiter=";")
        writer.writerow(headers)

        for r in rows:
            writer.writerow([_normalize_cell_value(r.get(c)) for c in cols])

        data = io.BytesIO(bio.getvalue().encode("utf-8-sig"))
        filename = f"pacientes_full_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        return send_file(data, as_attachment=True, download_name=filename, mimetype="text/csv")


# =============================================================================
# SECTION 9 · PDF (PRONTUÁRIO INDIVIDUAL) — LAYOUT TURBO
# =============================================================================

def _fmt(v: Any) -> str:
    s = "" if v is None else str(v).strip()
    return s if s else "—"

def _join_addr(p: dict) -> str:
    parts = []
    rua = (p.get("rua") or p.get("logradouro") or "").strip()
    num = (p.get("numero") or p.get("numero_casa") or "").strip()
    bairro = (p.get("bairro") or "").strip()
    cep = (p.get("cep") or "").strip()
    cid = (p.get("cidade") or p.get("municipio") or "").strip()
    uf = (p.get("uf") or "").strip()

    if rua: parts.append(rua)
    if num: parts.append(f"Nº {num}")
    if bairro: parts.append(bairro)
    if cep: parts.append(f"CEP {cep}")
    if cid: parts.append(cid)
    if uf: parts.append(uf)

    return " • ".join(parts) if parts else "—"

def _tags_human(p: dict) -> str:
    raw = p.get("comorbidades_json")
    keys = _json_list(raw)
    if not keys:
        return "—"
    mapa = dict(DEFAULT_TAGS)
    labels = [mapa.get(k, k) for k in keys]
    labels = [x for x in labels if str(x).strip()]
    return ", ".join(labels) if labels else "—"


@pacientes_bp.route("/exportar_prontuario_pdf/<int:id>")
def exportar_prontuario_pdf(id: int):
    """
    Exporta PDF do prontuário INDIVIDUAL do paciente (layout tipo prontuário/evolução).
    """
    ensure_pacientes_schema()

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm
        from reportlab.lib import colors
    except ImportError:
        return ("⚠️ Para exportar PDF, instale o pacote 'reportlab' (pip install reportlab).", 501)

    # ----- carrega paciente -----
    conn = conectar_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM pacientes WHERE id = ? LIMIT 1", (id,))
    row = cur.fetchone()
    conn.close()

    if not row:
        return ("Paciente não encontrado.", 404)

    p = dict(row)

    # compat do teu card
    if not (p.get("telefone") or "").strip():
        p["telefone"] = (p.get("telefone1") or "").strip()
    if not (p.get("nome_mae") or "").strip():
        p["nome_mae"] = (p.get("mae") or "").strip()
    if not (p.get("nome_pai") or "").strip():
        p["nome_pai"] = (p.get("pai") or "").strip()

    # idade calculada (se precisar)
    if p.get("idade") is None:
        p["idade"] = _calc_idade(p.get("nascimento"))

    # ----- agendamentos/enriquecimento -----
    ag_map = _get_primeiro_agendamento_por_paciente()
    info_ag = ag_map.get((p.get("nome") or "").strip().upper(), {}) if p.get("nome") else {}

    terapeuta = (info_ag.get("terapeuta_str") or "").strip()
    cbo_str   = (info_ag.get("cbo_str") or "").strip()
    ag_resumo = (info_ag.get("agenda_str") or "").strip()

    agds = _fetch_agendamentos_por_paciente(p.get("nome") or "")
    agds_upcoming = agds.get("agds_upcoming", [])

    # ----- PDF base -----
    bio = io.BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    W, H = A4

    margin = 14 * mm
    x0 = margin
    y = H - margin
    page_no = 1

    # cores suaves
    C_BORDER = colors.HexColor("#E5E7EB")
    C_SOFT   = colors.HexColor("#F8FAFC")
    C_SOFT2  = colors.HexColor("#F1F5F9")
    C_TEXT   = colors.HexColor("#0F172A")
    C_MUTED  = colors.HexColor("#475569")

    # ---------------- Helpers ----------------
    def new_page():
        nonlocal y, page_no
        c.showPage()
        page_no += 1
        y = H - margin
        draw_header()

    def ensure_space(mm_needed: float):
        nonlocal y
        if y < margin + (mm_needed * mm):
            new_page()

    def wrap_text(text: str, max_w: float, font="Helvetica", size=10) -> list[str]:
        c.setFont(font, size)
        text = (text or "").strip()
        if not text:
            return ["—"]
        words = text.split()
        lines = []
        curw = words[0]
        for w in words[1:]:
            test = curw + " " + w
            if c.stringWidth(test, font, size) <= max_w:
                curw = test
            else:
                lines.append(curw)
                curw = w
        lines.append(curw)
        return lines

    def draw_header():
        nonlocal y
        c.setFillColor(C_TEXT)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(x0, y, "Prontuário do Paciente")
        c.setFont("Helvetica", 9)
        c.setFillColor(C_MUTED)
        c.drawRightString(W - margin, y, f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        y -= 7 * mm

        c.setFillColor(C_TEXT)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(x0, y, _fmt(p.get("nome")))
        c.setFont("Helvetica", 9)
        c.setFillColor(C_MUTED)
        c.drawRightString(W - margin, y, f"Página {page_no}")
        y -= 6 * mm

        c.setStrokeColor(C_BORDER)
        c.setLineWidth(0.8)
        c.line(x0, y, W - margin, y)
        y -= 8 * mm

    def card_box(title: str, mm_h: float):
        nonlocal y
        ensure_space(mm_h + 10)
        w = W - 2 * margin
        h = mm_h * mm
        y_top = y

        c.setFillColor(C_SOFT)
        c.setStrokeColor(C_BORDER)
        c.setLineWidth(0.8)
        c.roundRect(x0, y_top - h, w, h, 10, stroke=1, fill=1)

        c.setFillColor(C_TEXT)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(x0 + 10, y_top - 16, title)

        y = y_top - 26
        return (x0 + 10, y, w - 20, h - 26)

    def draw_chip(x, y, label: str, value: Any, max_w: float):
        lab = (label or "").upper()
        val = _fmt(value)
        text = f"{lab}: {val}"

        c.setFont("Helvetica", 9)
        w = min(max_w, c.stringWidth(text, "Helvetica", 9) + 16)

        c.setFillColor(C_SOFT2)
        c.setStrokeColor(C_BORDER)
        c.roundRect(x, y - 12, w, 16, 8, stroke=1, fill=1)

        c.setFillColor(C_TEXT)
        c.drawString(x + 8, y, text)
        return w + 6

    def draw_kv(x, y, label: str, value: Any, col_w: float):
        c.setFillColor(C_MUTED)
        c.setFont("Helvetica", 8)
        c.drawString(x, y, (label or "").upper())
        y2 = y - 4.2 * mm

        c.setFillColor(C_TEXT)
        lines = wrap_text(_fmt(value), col_w, "Helvetica", 10)
        for ln in lines:
            c.setFont("Helvetica", 10)
            c.drawString(x, y2, ln)
            y2 -= 4.6 * mm

        return (y - y2) + (1.5 * mm)

    def draw_note_box(x, y, w, title: str, text: Any):
        c.setFillColor(C_SOFT2)
        c.setStrokeColor(C_BORDER)
        c.roundRect(x, y - 70, w, 70, 10, stroke=1, fill=1)

        c.setFillColor(C_TEXT)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(x + 10, y - 16, title)

        c.setFillColor(C_TEXT)
        lines = wrap_text(_fmt(text), w - 20, "Helvetica", 10)
        yy = y - 28
        for ln in lines[:6]:
            c.drawString(x + 10, yy, ln)
            yy -= 4.6 * mm
        if len(lines) > 6:
            c.setFont("Helvetica-Oblique", 9)
            c.setFillColor(C_MUTED)
            c.drawString(x + 10, yy, "… (texto cortado)")
        return 75

    # ---------------- Start ----------------
    draw_header()

    # 1) Identificação
    x, y_in, w_in, _ = card_box("Identificação", mm_h=56)

    col_gap = 12
    col_w = (w_in - col_gap) / 2
    left_x = x
    right_x = x + col_w + col_gap

    yy = y_in
    used = 0
    used += draw_chip(left_x, yy, "CPF", p.get("cpf"), col_w)
    used += draw_chip(left_x + used, yy, "CNS", p.get("cns"), col_w - used)

    used2 = 0
    used2 += draw_chip(right_x, yy, "Nascimento", p.get("nascimento"), col_w)
    used2 += draw_chip(right_x + used2, yy, "Idade", p.get("idade"), col_w - used2)

    yy -= 10 * mm
    h1 = draw_kv(left_x, yy, "Prontuário (código)", p.get("prontuario"), col_w)
    h2 = draw_kv(right_x, yy, "Sexo", p.get("sexo"), col_w)
    yy -= max(h1, h2)

    h1 = draw_kv(left_x, yy, "Telefone", p.get("telefone"), col_w)
    h2 = draw_kv(right_x, yy, "END (Prontuário físico)", p.get("end_prontuario"), col_w)
    yy -= max(h1, h2)

    y = (y_in - (56 * mm)) - 10

    # 2) Status e Classificação
    x, y_in, w_in, _ = card_box("Status e Classificação", mm_h=44)

    yy = y_in
    used = 0
    used += draw_chip(x, yy, "Status", p.get("status"), w_in)
    used += draw_chip(x + used, yy, "Modalidade", p.get("mod"), w_in - used)

    cid_combo = _fmt(p.get("cid"))
    if (p.get("cid2") or "").strip():
        cid_combo = f"{cid_combo} | CID2: {_fmt(p.get('cid2'))}"

    yy -= 10 * mm
    draw_kv(x, yy, "CID", cid_combo, w_in)

    y = (y_in - (44 * mm)) - 10

    # 3) Endereço e Família
    x, y_in, w_in, _ = card_box("Endereço e Família", mm_h=52)

    yy = y_in
    h1 = draw_kv(x, yy, "Endereço", _join_addr(p), w_in)
    yy -= h1

    col_gap = 12
    col_w = (w_in - col_gap) / 2
    left_x = x
    right_x = x + col_w + col_gap

    h1 = draw_kv(left_x, yy, "Nome da mãe", p.get("nome_mae"), col_w)
    h2 = draw_kv(right_x, yy, "Nome do pai", p.get("nome_pai"), col_w)
    yy -= max(h1, h2)

    y = (y_in - (52 * mm)) - 10

    # 4) Evolução / Observações
    x, y_in, w_in, _ = card_box("Evolução / Observações", mm_h=78)

    col_gap = 12
    col_w = (w_in - col_gap) / 2
    left_x = x
    right_x = x + col_w + col_gap
    yy = y_in + 6

    draw_note_box(left_x, yy, col_w, "Alergias", p.get("alergias"))
    draw_note_box(right_x, yy, col_w, "Aviso / Situação", p.get("aviso"))

    yy -= 78
    c.setFillColor(C_MUTED)
    c.setFont("Helvetica", 8)
    c.drawString(left_x, yy, "COMORBIDADES / PROJETOS")
    c.setFillColor(C_TEXT)
    c.setFont("Helvetica", 10)
    for ln in wrap_text(_tags_human(p), w_in, "Helvetica", 10)[:2]:
        yy -= 5 * mm
        c.drawString(left_x, yy, ln)

    y = (y_in - (78 * mm)) - 10

    # 5) Agenda / Terapias
    x, y_in, w_in, _ = card_box("Agenda / Terapias", mm_h=60)

    yy = y_in
    if terapeuta:
        draw_chip(x, yy, "Terapeuta(s)", terapeuta, w_in)

    if cbo_str:
        yy -= 10 * mm
        draw_chip(x, yy, "CBO(s)", cbo_str, w_in)

    yy -= 12 * mm
    draw_kv(x, yy, "Resumo do agendamento", _fmt(ag_resumo), w_in)

    y = (y_in - (60 * mm)) - 10

    # 6) Próximos agendamentos
    if agds_upcoming:
        x, y_in, w_in, _ = card_box("Próximos agendamentos", mm_h=80)
        yy = y_in

        c.setFont("Helvetica", 10)
        c.setFillColor(C_TEXT)

        max_items = 20
        count = 0
        for a in agds_upcoming:
            if count >= max_items:
                c.setFont("Helvetica-Oblique", 9)
                c.setFillColor(C_MUTED)
                c.drawString(x, yy, "… (lista cortada para manter o PDF leve)")
                yy -= 6 * mm
                break

            dia = _fmt(a.get("dia_semana"))
            data_br = _fmt(a.get("data_br"))
            hi = _fmt(a.get("hora_ini"))
            hf = (a.get("hora_fim") or "").strip()
            faixa = f"{hi}–{hf}" if hf else hi
            prof = _fmt(a.get("profissional"))

            linha = f"{dia} • {data_br} • {faixa} — {prof}"
            for ln in wrap_text(linha, w_in, "Helvetica", 10)[:2]:
                c.drawString(x, yy, ln)
                yy -= 5 * mm
            yy -= 1.5 * mm

            count += 1
            if yy < (y_in - (80 * mm) + 18):
                new_page()
                x, y_in, w_in, _ = card_box("Próximos agendamentos (continuação)", mm_h=80)
                yy = y_in

        y = (y_in - (80 * mm)) - 10

    c.showPage()
    c.save()
    bio.seek(0)

    nome_slug = re.sub(r"[^A-Za-z0-9]+", "_", (p.get("nome") or "paciente").strip()).strip("_")
    filename = f"prontuario_{nome_slug}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(bio, as_attachment=True, download_name=filename, mimetype="application/pdf")



# =============================================================================
# SECTION 10 · EXCLUSÃO
# =============================================================================

@pacientes_bp.route("/excluir/<int:id>", methods=["POST"])
def excluir_paciente(id):
    try:
        with conectar_db() as conn:
            try:
                conn.execute("PRAGMA foreign_keys = ON;")
            except Exception:
                pass

            cur = conn.cursor()
            cur.execute("SELECT 1 FROM pacientes WHERE id = ? LIMIT 1;", (id,))
            if not cur.fetchone():
                return ("Paciente não encontrado.", 404)

            cur.execute("DELETE FROM pacientes WHERE id = ?;", (id,))
            conn.commit()

        ref = request.referrer or url_for("pacientes.listar_pacientes")
        return redirect(ref)

    except sqlite3.IntegrityError as e:
        return (f"Não foi possível excluir (registro referenciado). Detalhes: {e}", 400)
    except Exception as e:
        return (f"Erro ao excluir paciente: {e}", 500)


# =============================================================================
# SECTION 11 · DIAGNÓSTICO
# =============================================================================

@pacientes_bp.route("/__ping")
def ping_pacientes():
    return "ok", 200
