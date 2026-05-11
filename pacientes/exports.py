# -*- coding: utf-8 -*-
from __future__ import annotations

import io
import json
import re
from datetime import datetime
from typing import Any

from flask import (
    send_file,
    request,
    session,
    abort,
)

from . import pacientes_bp
from .helpers import (
    fetch_pacientes_list,
    fetch_agendamentos_por_paciente,
    get_primeiro_agendamento_por_paciente,
    ensure_pacientes_schema,
    get_conn,
    fetchone_dict,
    calc_idade,
    fmt,
    join_addr,
    tags_human,
    table_columns,
)

try:
    from admin.modulos import require_permission
except Exception:
    def require_permission(modulo_codigo: str, acao: str = "ver"):
        def deco(fn):
            return fn
        return deco

try:
    from log import registrar_log, log_erro
except Exception:
    def registrar_log(*args, **kwargs): pass
    def log_erro(*args, **kwargs): pass


# =============================================================================
# MULTI-CLÍNICA
# =============================================================================

def _clinica_id_atual() -> int:
    clinica_id = session.get("clinica_id")

    if not clinica_id:
        abort(403)

    try:
        return int(clinica_id)
    except Exception:
        abort(403)


def _ensure_export_schema(conn):
    ensure_pacientes_schema(conn)

    cols = table_columns(conn, "pacientes")

    if "clinica_id" not in cols:
        cur = conn.cursor()
        cur.execute("""
            ALTER TABLE pacientes
            ADD COLUMN IF NOT EXISTS clinica_id INTEGER DEFAULT 1;
        """)
        conn.commit()

    cur = conn.cursor()

    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_pacientes_export_clinica
        ON pacientes(clinica_id);
    """)

    conn.commit()


def _buscar_paciente_clinica(conn, paciente_id: int):
    clinica_id = _clinica_id_atual()

    _ensure_export_schema(conn)

    cur = conn.cursor()

    cur.execute("""
        SELECT *
          FROM pacientes
         WHERE id = %s
           AND clinica_id = %s
         LIMIT 1;
    """, (paciente_id, clinica_id))

    return fetchone_dict(cur)


# =============================================================================
# HELPERS EXPORTAÇÃO
# =============================================================================

def export_header_order() -> list[str]:
    return [
        "id",
        "clinica_id",
        "prontuario",
        "nome",
        "nascimento",
        "idade",
        "sexo",
        "cpf",
        "cns",
        "telefone",
        "status",
        "mod",
        "cid",
        "cid2",

        # endereço
        "rua",
        "logradouro",
        "numero",
        "numero_casa",
        "bairro",
        "cep",
        "cidade",
        "municipio",
        "uf",

        # família
        "nome_mae",
        "mae",
        "nome_pai",
        "pai",

        # card / autosave
        "end_prontuario",
        "alergias",
        "aviso",
        "comorbidades_json",

        # derivados
        "terapeuta",
        "cbo",
        "ag_dia",
        "ag_hora_ini",
        "ag_hora_fim",
        "ag_resumo",
    ]


def pretty_header(col: str) -> str:
    mapa = {
        "id": "ID",
        "clinica_id": "Clínica ID",
        "prontuario": "Prontuário",
        "nome": "Nome",
        "nascimento": "Nascimento",
        "idade": "Idade",
        "sexo": "Sexo",
        "cpf": "CPF",
        "cns": "CNS",
        "telefone": "Telefone",
        "status": "Status",
        "mod": "Modalidade",
        "cid": "CID",
        "cid2": "CID 2",

        "rua": "Rua",
        "logradouro": "Logradouro",
        "numero": "Número",
        "numero_casa": "Número (casa)",
        "bairro": "Bairro",
        "cep": "CEP",
        "cidade": "Cidade",
        "municipio": "Município",
        "uf": "UF",

        "nome_mae": "Nome da mãe",
        "mae": "Mãe",
        "nome_pai": "Nome do pai",
        "pai": "Pai",

        "end_prontuario": "END (Prontuário)",
        "alergias": "Alergias",
        "aviso": "Aviso",
        "comorbidades_json": "Comorbidades",

        "terapeuta": "Terapeuta(s)",
        "cbo": "CBO(s)",
        "ag_dia": "Dia",
        "ag_hora_ini": "Hora início",
        "ag_hora_fim": "Hora fim",
        "ag_resumo": "Resumo agenda",
    }

    return mapa.get(col, col.replace("_", " ").strip().title())


def normalize_cell_value(v: Any) -> str:
    if v is None:
        return ""

    if isinstance(v, (dict, list)):
        try:
            return json.dumps(v, ensure_ascii=False)
        except Exception:
            return str(v)

    return str(v)


# =============================================================================
# EXPORT XLS/CSV
# =============================================================================

@pacientes_bp.route("/exportar_xls")
@require_permission("pacientes", "exportar")
def exportar_xls():
    clinica_id = _clinica_id_atual()

    try:
        rows = fetch_pacientes_list(request.args)

        keys_all: set[str] = set()

        for r in rows:
            if isinstance(r, dict):
                keys_all.update(r.keys())

        preferred = export_header_order()

        cols = [c for c in preferred if c in keys_all]

        resto = sorted([c for c in keys_all if c not in cols])

        cols.extend(resto)

        headers = [pretty_header(c) for c in cols]

        registrar_log(
            modulo="pacientes",
            acao="exportar",
            entidade="pacientes",
            descricao="Exportou pacientes XLS/CSV.",
            detalhes={
                "clinica_id": clinica_id,
                "total": len(rows),
                "filtros": dict(request.args),
            },
        )

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Pacientes"

            ws.append(headers)

            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

            for r in rows:
                ws.append([
                    normalize_cell_value(r.get(c))
                    for c in cols
                ])

            for idx, _col_name in enumerate(cols, start=1):
                letter = get_column_letter(idx)

                max_len = len(headers[idx - 1])

                for cell in ws[letter]:
                    if cell.value is None:
                        continue

                    max_len = max(max_len, len(str(cell.value)))

                ws.column_dimensions[letter].width = min(max_len + 2, 60)

            ws.freeze_panes = "A2"

            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)

            filename = (
                f"pacientes_"
                f"{clinica_id}_"
                f"{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            )

            return send_file(
                bio,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except ImportError:
            import csv

            bio_txt = io.StringIO()

            writer = csv.writer(bio_txt, delimiter=";")

            writer.writerow(headers)

            for r in rows:
                writer.writerow([
                    normalize_cell_value(r.get(c))
                    for c in cols
                ])

            data = io.BytesIO(
                bio_txt.getvalue().encode("utf-8-sig")
            )

            filename = (
                f"pacientes_"
                f"{clinica_id}_"
                f"{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
            )

            return send_file(
                data,
                as_attachment=True,
                download_name=filename,
                mimetype="text/csv",
            )

    except Exception as e:
        log_erro(
            "pacientes",
            e,
            entidade="pacientes",
            descricao="Erro ao exportar XLS/CSV.",
            detalhes={
                "clinica_id": clinica_id,
                "filtros": dict(request.args),
            },
        )

        return f"Erro ao exportar: {e}", 500


# =============================================================================
# PDF PRONTUÁRIO COM TIMBRE
# =============================================================================

def _row_to_dict(row, cur=None):
    if not row:
        return None
    if isinstance(row, dict):
        return dict(row)
    cols = [d[0] for d in cur.description] if cur and cur.description else []
    return dict(zip(cols, row))


def _normalizar_img_bin(v):
    try:
        from clinica_config.timbre import normalizar_binario_imagem
        return normalizar_binario_imagem(v)
    except Exception:
        if isinstance(v, memoryview):
            return v.tobytes()
        if isinstance(v, bytearray):
            return bytes(v)
        if isinstance(v, bytes):
            return v
        return None


def _buscar_timbre_pdf(conn, clinica_id: int) -> dict:
    cur = conn.cursor()
    cur.execute("""
        SELECT *
          FROM clinica_configuracoes
         WHERE clinica_id = %s
         ORDER BY id DESC
         LIMIT 1;
    """, (clinica_id,))

    cfg = _row_to_dict(cur.fetchone(), cur) or {}

    return {
        "cabecalho_texto": cfg.get("cabecalho_texto") or "",
        "rodape_texto": cfg.get("rodape_texto") or "",
        "cor_listra_topo": cfg.get("cor_listra_topo") or "#0f766e",
        "mostrar_linha_cabecalho": cfg.get("mostrar_linha_cabecalho", True),
        "mostrar_linha_rodape": cfg.get("mostrar_linha_rodape", True),
        "logo": _normalizar_img_bin(cfg.get("logo_bin")),
        "cabecalho": _normalizar_img_bin(cfg.get("cabecalho_img_bin")),
        "rodape1": _normalizar_img_bin(cfg.get("rodape_img_bin")),
        "rodape2": _normalizar_img_bin(cfg.get("rodape_img_2_bin")),
        "rodape3": _normalizar_img_bin(cfg.get("rodape_img_3_bin")),
    }


def _draw_img_fit(c, img_bytes, x, y, w, h):
    if not img_bytes:
        return

    try:
        import io
        from reportlab.lib.utils import ImageReader

        img = ImageReader(io.BytesIO(img_bytes))
        iw, ih = img.getSize()
        scale = min(w / iw, h / ih)
        nw, nh = iw * scale, ih * scale
        c.drawImage(img, x + (w - nw) / 2, y + (h - nh) / 2, nw, nh, mask="auto")
    except Exception:
        pass


def _pdf_wrap(c, text, max_w, font="Helvetica", size=9):
    text = fmt(text)

    if text == "—":
        return ["—"]

    c.setFont(font, size)

    words = text.split()
    lines = []
    current = ""

    for word in words:
        test = f"{current} {word}".strip()

        if c.stringWidth(test, font, size) <= max_w:
            current = test
        else:
            if current:
                lines.append(current)
            current = word

    if current:
        lines.append(current)

    return lines or ["—"]


def _draw_timbre(c, W, H, timbre, pagina=1):
    from reportlab.lib.units import mm
    from reportlab.lib import colors

    margin = 14 * mm

    try:
        cor = colors.HexColor(timbre.get("cor_listra_topo") or "#0f766e")
    except Exception:
        cor = colors.HexColor("#0f766e")

    c.setFillColor(cor)
    c.rect(0, H - 5 * mm, W, 5 * mm, stroke=0, fill=1)

    if timbre.get("cabecalho"):
        _draw_img_fit(
            c,
            timbre["cabecalho"],
            margin,
            H - 34 * mm,
            W - 2 * margin,
            25 * mm,
        )
        header_bottom = H - 39 * mm
    else:
        if timbre.get("logo"):
            _draw_img_fit(
                c,
                timbre["logo"],
                margin,
                H - 32 * mm,
                30 * mm,
                23 * mm,
            )

        texto = timbre.get("cabecalho_texto") or "PRONTUÁRIO DO PACIENTE"

        c.setFillColor(colors.HexColor("#0f172a"))
        c.setFont("Helvetica-Bold", 10)

        y = H - 16 * mm
        for ln in _pdf_wrap(c, texto, W - 58 * mm, "Helvetica-Bold", 10)[:3]:
            c.drawString(margin + 36 * mm, y, ln)
            y -= 4.5 * mm

        header_bottom = H - 36 * mm

    if timbre.get("mostrar_linha_cabecalho", True):
        c.setStrokeColor(colors.HexColor("#e2e8f0"))
        c.line(margin, header_bottom, W - margin, header_bottom)

    footer_top = 22 * mm

    if timbre.get("mostrar_linha_rodape", True):
        c.setStrokeColor(colors.HexColor("#e2e8f0"))
        c.line(margin, footer_top + 5 * mm, W - margin, footer_top + 5 * mm)

    rodapes = [
        timbre.get("rodape1"),
        timbre.get("rodape2"),
        timbre.get("rodape3"),
    ]

    imgs = [r for r in rodapes if r]

    if imgs:
        img_w = (W - 2 * margin) / len(imgs)

        for idx, img in enumerate(imgs):
            _draw_img_fit(
                c,
                img,
                margin + idx * img_w,
                6 * mm,
                img_w - 3 * mm,
                15 * mm,
            )
    else:
        c.setFont("Helvetica", 7.5)
        c.setFillColor(colors.HexColor("#64748b"))

        texto = timbre.get("rodape_texto") or f"Gerado pelo SGD Saúde em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        c.drawCentredString(W / 2, 12 * mm, texto[:170])

    c.setFont("Helvetica", 7)
    c.setFillColor(colors.HexColor("#94a3b8"))
    c.drawRightString(W - margin, 6 * mm, f"Página {pagina}")

    return header_bottom - 8 * mm, footer_top + 11 * mm


def _draw_section_box(c, x, y, w, title, items, cols=2):
    from reportlab.lib.units import mm
    from reportlab.lib import colors

    title_h = 8 * mm
    row_h = 12.5 * mm
    rows = max(1, (len(items) + cols - 1) // cols)
    h = title_h + rows * row_h + 6 * mm

    c.setFillColor(colors.white)
    c.setStrokeColor(colors.HexColor("#e2e8f0"))
    c.roundRect(x, y - h, w, h, 8, stroke=1, fill=1)

    c.setFillColor(colors.HexColor("#0f172a"))
    c.setFont("Helvetica-Bold", 10)
    c.drawString(x + 5 * mm, y - 5.5 * mm, title)

    col_w = (w - 10 * mm) / cols
    start_y = y - title_h - 4 * mm

    for idx, (label, value) in enumerate(items):
        cx = x + 5 * mm + (idx % cols) * col_w
        cy = start_y - (idx // cols) * row_h

        c.setFillColor(colors.HexColor("#64748b"))
        c.setFont("Helvetica-Bold", 7)
        c.drawString(cx, cy, str(label).upper())

        c.setFillColor(colors.HexColor("#0f172a"))
        c.setFont("Helvetica", 9)

        yy = cy - 4 * mm
        for ln in _pdf_wrap(c, value, col_w - 4 * mm, "Helvetica", 9)[:2]:
            c.drawString(cx, yy, ln)
            yy -= 3.8 * mm

    return y - h - 5 * mm


@pacientes_bp.route("/exportar_prontuario_pdf/<int:id>")
@require_permission("pacientes", "exportar")
def exportar_prontuario_pdf(id: int):
    clinica_id = _clinica_id_atual()

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm
        from reportlab.lib import colors
    except ImportError:
        return "⚠️ Instale reportlab: pip install reportlab", 501

    try:
        with get_conn() as conn:
            paciente = _buscar_paciente_clinica(conn, id)
            timbre = _buscar_timbre_pdf(conn, clinica_id)

        if not paciente:
            return "Paciente não encontrado nesta clínica.", 404

        p = dict(paciente)

        if not (p.get("telefone") or "").strip():
            p["telefone"] = (p.get("telefone1") or "").strip()

        if not (p.get("nome_mae") or "").strip():
            p["nome_mae"] = (p.get("mae") or "").strip()

        if not (p.get("nome_pai") or "").strip():
            p["nome_pai"] = (p.get("pai") or "").strip()

        if p.get("idade") is None:
            p["idade"] = calc_idade(p.get("nascimento"))

        try:
            ag_map = get_primeiro_agendamento_por_paciente()
            info_ag = ag_map.get((p.get("nome") or "").strip().upper(), {})
        except Exception:
            info_ag = {}

        try:
            agds = fetch_agendamentos_por_paciente(
                p.get("nome") or "",
                clinica_id=clinica_id,
            )
        except TypeError:
            agds = fetch_agendamentos_por_paciente(p.get("nome") or "")

        agds_upcoming = agds.get("agds_upcoming", [])

        bio = io.BytesIO()
        c = canvas.Canvas(bio, pagesize=A4)

        W, H = A4
        margin = 14 * mm
        box_w = W - 2 * margin
        page = 1

        def iniciar_pagina():
            return _draw_timbre(c, W, H, timbre, page)

        y, content_bottom = iniciar_pagina()

        def nova_pagina():
            nonlocal page, y, content_bottom
            c.showPage()
            page += 1
            y, content_bottom = iniciar_pagina()

        def precisa(altura_mm):
            if y - altura_mm * mm < content_bottom:
                nova_pagina()

        c.setFillColor(colors.HexColor("#0f172a"))
        c.setFont("Helvetica-Bold", 15)
        c.drawString(margin, y, "Prontuário do Paciente")

        c.setFont("Helvetica", 8)
        c.setFillColor(colors.HexColor("#64748b"))
        c.drawRightString(W - margin, y, f"Emitido em {datetime.now().strftime('%d/%m/%Y %H:%M')}")

        y -= 8 * mm

        c.setFillColor(colors.HexColor("#111827"))
        c.setFont("Helvetica-Bold", 13)

        for ln in _pdf_wrap(c, p.get("nome"), box_w, "Helvetica-Bold", 13)[:2]:
            c.drawString(margin, y, ln)
            y -= 6 * mm

        y -= 2 * mm

        precisa(42)
        y = _draw_section_box(c, margin, y, box_w, "Identificação", [
            ("Prontuário", p.get("prontuario")),
            ("Modalidade", p.get("mod")),
            ("Status", p.get("status")),
            ("Nascimento", p.get("nascimento")),
            ("Idade", p.get("idade")),
            ("Sexo", p.get("sexo")),
            ("CPF", p.get("cpf")),
            ("CNS", p.get("cns")),
        ], cols=4)

        precisa(42)
        y = _draw_section_box(c, margin, y, box_w, "Contato e Endereço", [
            ("Telefone", p.get("telefone")),
            ("Telefone 2", p.get("telefone2")),
            ("Telefone 3", p.get("telefone3")),
            ("E-mail", p.get("email")),
            ("Logradouro", p.get("logradouro") or p.get("rua")),
            ("Número", p.get("numero") or p.get("numero_casa")),
            ("Bairro", p.get("bairro")),
            ("Cidade / UF", f"{fmt(p.get('municipio') or p.get('cidade'))} / {fmt(p.get('uf'))}"),
            ("CEP", p.get("cep")),
        ], cols=3)

        precisa(34)
        y = _draw_section_box(c, margin, y, box_w, "Documentos e Dados Sociais", [
            ("RG", p.get("rg")),
            ("Órgão RG", p.get("orgao_rg")),
            ("Estado civil", p.get("estado_civil")),
            ("NIS", p.get("nis")),
            ("Raça/Cor", p.get("raca")),
        ], cols=3)

        precisa(44)
        y = _draw_section_box(c, margin, y, box_w, "Família e Responsável", [
            ("Mãe", p.get("nome_mae") or p.get("mae")),
            ("CPF mãe", p.get("cpf_mae")),
            ("RG mãe", p.get("rg_mae")),
            ("Pai", p.get("nome_pai") or p.get("pai")),
            ("CPF pai", p.get("cpf_pai")),
            ("RG pai", p.get("rg_pai")),
            ("Responsável", p.get("responsavel")),
            ("CPF responsável", p.get("cpf_responsavel")),
            ("RG responsável", p.get("rg_responsavel")),
        ], cols=3)

        precisa(44)
        y = _draw_section_box(c, margin, y, box_w, "Dados Clínicos", [
            ("CID principal", p.get("cid")),
            ("CID secundário", p.get("cid2")),
            ("Alergias", p.get("alergias")),
            ("Comorbidades", tags_human(p)),
            ("Aviso / Situação", p.get("aviso")),
            ("END prontuário físico", p.get("end_prontuario")),
        ], cols=2)

        precisa(34)
        y = _draw_section_box(c, margin, y, box_w, "Agenda / Terapias", [
            ("Terapeuta(s)", info_ag.get("terapeuta_str") or p.get("terapeuta")),
            ("CBO(s)", info_ag.get("cbo_str") or p.get("cbo")),
            ("Resumo", info_ag.get("agenda_str") or p.get("ag_resumo")),
        ], cols=1)

        if agds_upcoming:
            precisa(38)

            c.setFillColor(colors.HexColor("#0f172a"))
            c.setFont("Helvetica-Bold", 10)
            c.drawString(margin, y, "Próximos agendamentos")
            y -= 6 * mm

            c.setFont("Helvetica", 9)
            c.setFillColor(colors.HexColor("#111827"))

            for a in agds_upcoming[:25]:
                precisa(10)

                linha = (
                    f"{fmt(a.get('dia_semana'))} • "
                    f"{fmt(a.get('data_br'))} • "
                    f"{fmt(a.get('hora_ini'))}"
                    f"{'–' + a.get('hora_fim') if a.get('hora_fim') else ''}"
                    f" — {fmt(a.get('profissional'))}"
                )

                for ln in _pdf_wrap(c, linha, box_w, "Helvetica", 9)[:2]:
                    c.drawString(margin, y, ln)
                    y -= 4.5 * mm

                y -= 1 * mm

        c.save()
        bio.seek(0)

        registrar_log(
            modulo="pacientes",
            acao="exportar",
            entidade="pacientes",
            entidade_id=id,
            descricao="Exportou prontuário PDF com timbre.",
            detalhes={
                "clinica_id": clinica_id,
                "paciente_id": id,
                "nome": p.get("nome"),
                "com_timbre": True,
            },
        )

        nome_slug = re.sub(
            r"[^A-Za-z0-9]+",
            "_",
            (p.get("nome") or "paciente").strip(),
        ).strip("_")

        filename = f"prontuario_{nome_slug}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/pdf",
        )

    except Exception as e:
        log_erro(
            "pacientes",
            e,
            entidade="pacientes",
            entidade_id=id,
            descricao="Erro ao exportar prontuário PDF com timbre.",
            detalhes={
                "clinica_id": clinica_id,
                "paciente_id": id,
            },
        )

        return f"Erro ao gerar PDF: {e}", 500