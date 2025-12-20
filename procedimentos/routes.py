# sgd/procedimentos/routes.py
from __future__ import annotations

import io
import sqlite3
from datetime import datetime

from flask import (
    request, jsonify, render_template, flash, redirect, url_for, send_file
)
from openpyxl import load_workbook, Workbook

from . import procedimentos_bp
from db import conectar_db


# ============================================================
# DICIONÁRIO DE PROFISSIONAIS ↔ CBO (referência / UI)
# ============================================================

CBO_GRUPOS: list[dict] = [
    {"id": 1, "nome": "Serviço social", "codigos": ["251605"]},
    {"id": 2, "nome": "Enfermagem", "codigos": ["223505"]},
    {"id": 3, "nome": "Psicólogos", "codigos": ["251510"]},
    {"id": 4, "nome": "Terapeuta Ocupacional", "codigos": ["223905"]},
    {"id": 5, "nome": "Fonoaudiólogos", "codigos": ["223810"]},
    {"id": 6, "nome": "Fisioterapeutas", "codigos": ["223605"]},
    {"id": 7, "nome": "Pedagogos", "codigos": ["239425", "239415"]},
    {"id": 8, "nome": "Nutricionista", "codigos": ["223710"]},
    {"id": 9, "nome": "Neurologista", "codigos": ["225112"]},
    {"id": 10, "nome": "Clínico", "codigos": ["225125"]},
    {"id": 11, "nome": "Psiquiatra", "codigos": ["225133"]},
    {"id": 12, "nome": "Ortopedista", "codigos": ["225270"]},
    {"id": 13, "nome": "Otorrino", "codigos": ["225275"]},
]

CBO_POR_CODIGO: dict[str, str] = {}
for g in CBO_GRUPOS:
    for cod in g["codigos"]:
        CBO_POR_CODIGO[cod] = g["nome"]


# ============================================================
# (Opcional) dicionário em memória — fallback
# ============================================================

PROCEDIMENTOS_POR_CBO: dict[str, list[dict]] = {
    "223605": [
        {"codigo": "0301010030", "descricao": "Avaliação fisioterapêutica", "cid": None, "idade_min": None, "idade_max": None, "sexo": None},
        {"codigo": "0301010048", "descricao": "Atendimento fisioterapêutico individual", "cid": None, "idade_min": None, "idade_max": None, "sexo": None},
    ],
    "251510": [
        {"codigo": "0301010021", "descricao": "Atendimento psicológico individual", "cid": None, "idade_min": None, "idade_max": None, "sexo": None},
    ],
    "223710": [
        {"codigo": "0301010060", "descricao": "Atendimento nutricional individual", "cid": None, "idade_min": None, "idade_max": None, "sexo": None},
    ],
}


# ============================================================
# Helpers / Normalização
# ============================================================

COLS_ESPERADAS = [
    "PA_COD",
    "PA_CID",
    "PA_DESCRICAO",
    "PA_IDADEMN",
    "PA_IDADEMX",
    "PA_SEXO",
    "PA_CBO",
    "PA_CBO_NAME",
]


def _norm_txt(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _norm_int(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(str(v).strip()))
    except Exception:
        return None


def _norm_sexo(v) -> str | None:
    s = _norm_txt(v).upper()
    if not s:
        return None
    s = s[0]  # M/F/A
    if s not in ("M", "F", "A"):
        return None
    return s


def _now_sql() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _row_to_dict(r):
    if hasattr(r, "keys"):
        return {k: r[k] for k in r.keys()}
    return r


def _norm_regra_fields(it: dict) -> dict:
    """
    ⭐ CRÍTICO:
    Para o UNIQUE funcionar no SQLite, NÃO PODE deixar NULL em campos da regra.
    Senão: NULL != NULL e você cria duplicatas “fantasma”.
    """
    it = dict(it)

    it["pa_cid"] = _norm_txt(it.get("pa_cid"))  # '' se vazio
    it["pa_sexo"] = _norm_txt(it.get("pa_sexo"))  # '' se vazio

    mn = it.get("pa_idademn")
    mx = it.get("pa_idademx")

    it["pa_idademn"] = -1 if (mn is None or str(mn).strip() == "") else int(mn)
    it["pa_idademx"] = -1 if (mx is None or str(mx).strip() == "") else int(mx)

    # garante strings
    it["pa_cod"] = _norm_txt(it.get("pa_cod"))
    it["pa_descricao"] = _norm_txt(it.get("pa_descricao"))
    it["pa_cbo"] = _norm_txt(it.get("pa_cbo"))
    it["pa_cbo_name"] = _norm_txt(it.get("pa_cbo_name")) or None

    return it


# ============================================================
# SQLite schema
# ============================================================

def ensure_procedimentos_schema(conn):
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS procedimentos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,

            pa_cod        TEXT NOT NULL,
            pa_cid        TEXT NOT NULL DEFAULT '',
            pa_descricao  TEXT NOT NULL,
            pa_idademn    INTEGER NOT NULL DEFAULT -1,
            pa_idademx    INTEGER NOT NULL DEFAULT -1,
            pa_sexo       TEXT NOT NULL DEFAULT '',
            pa_cbo        TEXT NOT NULL,
            pa_cbo_name   TEXT,

            ativo         INTEGER NOT NULL DEFAULT 1,
            criado_em     TEXT NOT NULL DEFAULT (datetime('now')),
            atualizado_em TEXT
        )
    """)

    # remove índices antigos (se existirem)
    cur.execute("DROP INDEX IF EXISTS ux_procedimentos_cod_cbo")
    cur.execute("DROP INDEX IF EXISTS ux_procedimentos_regra")

    # ✅ índice único alinhado com a REGRA COMPLETA
    cur.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS ux_procedimentos_regra
        ON procedimentos (
            pa_cod,
            pa_cbo,
            pa_cid,
            pa_sexo,
            pa_idademn,
            pa_idademx
        )
    """)

    conn.commit()


# ============================================================
# Leitura Excel (XLSX)
# ============================================================

def _ler_xlsx_procedimentos(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    header = [(_norm_txt(c.value).upper()) for c in ws[1]]
    header_map = {name: idx for idx, name in enumerate(header)}

    faltando = [c for c in COLS_ESPERADAS if c not in header_map]
    if faltando:
        return [], [f"Colunas ausentes no Excel: {', '.join(faltando)}"]

    rows: list[dict] = []
    erros: list[str] = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def get(col: str):
            idx = header_map[col]
            return row[idx] if idx < len(row) else None

        pa_cod = _norm_txt(get("PA_COD"))
        pa_desc = _norm_txt(get("PA_DESCRICAO"))
        pa_cbo = _norm_txt(get("PA_CBO"))

        if not pa_cod and not pa_desc and not pa_cbo:
            continue

        if not pa_cod or not pa_desc or not pa_cbo:
            erros.append(f"Linha {i}: PA_COD/PA_DESCRICAO/PA_CBO são obrigatórios.")
            continue

        it = {
            "pa_cod": pa_cod,
            "pa_cid": _norm_txt(get("PA_CID")) or "",
            "pa_descricao": pa_desc,
            "pa_idademn": _norm_int(get("PA_IDADEMN")),
            "pa_idademx": _norm_int(get("PA_IDADEMX")),
            "pa_sexo": _norm_sexo(get("PA_SEXO")) or "",
            "pa_cbo": pa_cbo,
            "pa_cbo_name": _norm_txt(get("PA_CBO_NAME")) or None,
        }

        rows.append(_norm_regra_fields(it))

    return rows, erros


# ============================================================
# CRUD helpers (form -> dict)
# ============================================================

def _ler_form_procedimento() -> tuple[dict | None, str | None]:
    pa_cod = _norm_txt(request.form.get("pa_cod"))
    pa_desc = _norm_txt(request.form.get("pa_descricao"))
    pa_cbo = _norm_txt(request.form.get("pa_cbo"))

    if not pa_cod or not pa_desc or not pa_cbo:
        return None, "PA_COD, PA_DESCRICAO e PA_CBO são obrigatórios."

    data = {
        "pa_cod": pa_cod,
        "pa_cid": _norm_txt(request.form.get("pa_cid")) or "",
        "pa_descricao": pa_desc,
        "pa_idademn": _norm_int(request.form.get("pa_idademn")),
        "pa_idademx": _norm_int(request.form.get("pa_idademx")),
        "pa_sexo": _norm_sexo(request.form.get("pa_sexo")) or "",
        "pa_cbo": pa_cbo,
        "pa_cbo_name": _norm_txt(request.form.get("pa_cbo_name")) or None,
    }
    return _norm_regra_fields(data), None


# ============================================================
# UPSERT por regra completa (casa com o índice UNIQUE)
# ============================================================

def _upsert_regra(cur, it: dict, now: str) -> tuple[int, int]:
    """
    UPSERT por regra completa:
    - Tenta inserir (INSERT OR IGNORE)
    - Se ignorou, faz UPDATE por mesma regra completa
    Retorna (inseridos, atualizados)
    """
    it = _norm_regra_fields(it)

    insert_sql = """
        INSERT OR IGNORE INTO procedimentos (
          pa_cod, pa_cid, pa_descricao, pa_idademn, pa_idademx, pa_sexo,
          pa_cbo, pa_cbo_name, ativo, atualizado_em
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
    """

    update_sql = """
        UPDATE procedimentos
           SET pa_descricao  = ?,
               pa_cbo_name   = ?,
               ativo         = 1,
               atualizado_em = ?
         WHERE pa_cod = ?
           AND pa_cbo = ?
           AND pa_cid = ?
           AND pa_sexo = ?
           AND pa_idademn = ?
           AND pa_idademx = ?
    """

    cur.execute(
        insert_sql,
        (
            it["pa_cod"],
            it["pa_cid"],
            it["pa_descricao"],
            it["pa_idademn"],
            it["pa_idademx"],
            it["pa_sexo"],
            it["pa_cbo"],
            it["pa_cbo_name"],
            now,
        ),
    )

    if cur.rowcount == 1:
        return 1, 0

    cur.execute(
        update_sql,
        (
            it["pa_descricao"],
            it["pa_cbo_name"],
            now,
            it["pa_cod"],
            it["pa_cbo"],
            it["pa_cid"],
            it["pa_sexo"],
            it["pa_idademn"],
            it["pa_idademx"],
        ),
    )
    return 0, 1


# ============================================================
# Queries
# ============================================================

def _listar_procedimentos_db(conn, limit: int = 200):
    ensure_procedimentos_schema(conn)
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, pa_cod, pa_cid, pa_descricao, pa_idademn, pa_idademx, pa_sexo, pa_cbo, pa_cbo_name
          FROM procedimentos
         WHERE ativo = 1
         ORDER BY pa_cbo ASC, pa_descricao COLLATE NOCASE ASC
         LIMIT ?
        """,
        (int(limit),),
    )
    rows = cur.fetchall()

    out: list[dict] = []
    for r in rows:
        if hasattr(r, "keys"):
            out.append({k: r[k] for k in r.keys()})
        else:
            out.append(
                {
                    "id": r[0],
                    "pa_cod": r[1],
                    "pa_cid": r[2],
                    "pa_descricao": r[3],
                    "pa_idademn": r[4],
                    "pa_idademx": r[5],
                    "pa_sexo": r[6],
                    "pa_cbo": r[7],
                    "pa_cbo_name": r[8],
                }
            )
    return out


def filtrar_procedimentos(cbo: str, cid: str | None = None, idade: int | None = None, sexo: str | None = None) -> list[dict]:
    if not cbo:
        return []

    conn = conectar_db()
    ensure_procedimentos_schema(conn)
    cur = conn.cursor()

    where = ["pa_cbo = ?", "ativo = 1"]
    params: list = [cbo]

    if cid:
        where.append("(pa_cid = '' OR pa_cid = ?)")
        params.append(cid.strip())

    if idade is not None:
        where.append("(pa_idademn = -1 OR pa_idademn <= ?)")
        where.append("(pa_idademx = -1 OR pa_idademx >= ?)")
        params.extend([idade, idade])

    if sexo:
        sx = _norm_sexo(sexo)
        if sx:
            where.append("(pa_sexo = '' OR pa_sexo = 'A' OR pa_sexo = ?)")
            params.append(sx)

    sql = f"""
        SELECT pa_cod, pa_cid, pa_descricao, pa_idademn, pa_idademx, pa_sexo, pa_cbo, pa_cbo_name
          FROM procedimentos
         WHERE {' AND '.join(where)}
         ORDER BY pa_descricao COLLATE NOCASE ASC
    """
    cur.execute(sql, params)
    rows = cur.fetchall()

    if rows:
        out: list[dict] = []
        for r in rows:
            if hasattr(r, "keys"):
                out.append({k: r[k] for k in r.keys()})
            else:
                out.append(
                    {
                        "pa_cod": r[0],
                        "pa_cid": r[1],
                        "pa_descricao": r[2],
                        "pa_idademn": r[3],
                        "pa_idademx": r[4],
                        "pa_sexo": r[5],
                        "pa_cbo": r[6],
                        "pa_cbo_name": r[7],
                    }
                )
        return out

    # fallback dicionário
    todos = PROCEDIMENTOS_POR_CBO.get(cbo, [])
    out2: list[dict] = []
    for p in todos:
        if cid and p.get("cid") and p["cid"] != cid:
            continue
        if idade is not None:
            if p.get("idade_min") is not None and idade < p["idade_min"]:
                continue
            if p.get("idade_max") is not None and idade > p["idade_max"]:
                continue
        out2.append(
            {
                "pa_cod": p["codigo"],
                "pa_cid": p.get("cid") or "",
                "pa_descricao": p["descricao"],
                "pa_idademn": p.get("idade_min") if p.get("idade_min") is not None else -1,
                "pa_idademx": p.get("idade_max") if p.get("idade_max") is not None else -1,
                "pa_sexo": p.get("sexo") or "",
                "pa_cbo": cbo,
                "pa_cbo_name": CBO_POR_CODIGO.get(cbo),
            }
        )
    return out2


# ============================================================
# Rotas
# ============================================================

@procedimentos_bp.get("/")
def pagina_base_procedimentos():
    conn = conectar_db()
    procedimentos_preview = _listar_procedimentos_db(conn, limit=50)

    return render_template(
        "procedimentos.html",
        cbo_grupos=CBO_GRUPOS,
        procedimentos=procedimentos_preview,
    )


@procedimentos_bp.get("/api/sugeridos")
def api_procedimentos_sugeridos():
    cbo = (request.args.get("cbo") or "").strip()
    cid = (request.args.get("cid") or "").strip() or None
    sexo = (request.args.get("sexo") or "").strip() or None

    idade_raw = request.args.get("idade")
    idade: int | None = None
    if idade_raw and str(idade_raw).isdigit():
        idade = int(idade_raw)

    itens = filtrar_procedimentos(cbo=cbo, cid=cid, idade=idade, sexo=sexo)

    return jsonify(
        {
            "ok": True,
            "cbo": cbo,
            "cbo_nome": CBO_POR_CODIGO.get(cbo),
            "cid": cid,
            "idade": idade,
            "sexo": _norm_sexo(sexo) if sexo else None,
            "total": len(itens),
            "items": itens,
            "cols": COLS_ESPERADAS,
        }
    )


@procedimentos_bp.post("/criar")
def criar_procedimento():
    data, err = _ler_form_procedimento()
    if err:
        flash(err, "error")
        return redirect(url_for("procedimentos.pagina_base_procedimentos"))

    conn = conectar_db()
    ensure_procedimentos_schema(conn)
    cur = conn.cursor()

    now = _now_sql()
    try:
        _upsert_regra(cur, data, now)
        conn.commit()
    except sqlite3.IntegrityError:
        conn.rollback()
        flash("Conflito: já existe um procedimento com a MESMA regra (código+cbo+cid+sexo+idades).", "error")
        return redirect(url_for("procedimentos.pagina_base_procedimentos"))

    flash("Procedimento salvo com sucesso.", "success")
    return redirect(url_for("procedimentos.pagina_base_procedimentos"))


@procedimentos_bp.post("/<int:proc_id>/editar")
def editar_procedimento(proc_id: int):
    data, err = _ler_form_procedimento()
    if err:
        flash(err, "error")
        return redirect(url_for("procedimentos.pagina_base_procedimentos"))

    conn = conectar_db()
    ensure_procedimentos_schema(conn)
    cur = conn.cursor()

    now = _now_sql()

    try:
        cur.execute(
            """
            UPDATE procedimentos
               SET pa_cod = ?,
                   pa_cid = ?,
                   pa_descricao = ?,
                   pa_idademn = ?,
                   pa_idademx = ?,
                   pa_sexo = ?,
                   pa_cbo = ?,
                   pa_cbo_name = ?,
                   atualizado_em = ?
             WHERE id = ?
            """,
            (
                data["pa_cod"],
                data["pa_cid"],
                data["pa_descricao"],
                data["pa_idademn"],
                data["pa_idademx"],
                data["pa_sexo"],
                data["pa_cbo"],
                data["pa_cbo_name"],
                now,
                proc_id,
            ),
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.rollback()
        flash("Conflito: essa edição cria uma regra duplicada (mesmo código+cbo+cid+sexo+idades).", "error")
        return redirect(url_for("procedimentos.pagina_base_procedimentos"))

    flash("Procedimento atualizado.", "success")
    return redirect(url_for("procedimentos.pagina_base_procedimentos"))


@procedimentos_bp.post("/<int:proc_id>/remover")
def remover_procedimento(proc_id: int):
    conn = conectar_db()
    ensure_procedimentos_schema(conn)
    cur = conn.cursor()
    cur.execute(
        "UPDATE procedimentos SET ativo = 0, atualizado_em = ? WHERE id = ?",
        (_now_sql(), proc_id),
    )
    conn.commit()

    flash("Procedimento removido (desativado).", "success")
    return redirect(url_for("procedimentos.pagina_base_procedimentos"))


@procedimentos_bp.post("/importar")
def importar_planilha_procedimentos():
    f = request.files.get("file")
    if not f or not f.filename:
        flash("Envie um arquivo .xlsx no campo 'file'.", "error")
        return redirect(url_for("procedimentos.pagina_base_procedimentos"))

    if not f.filename.lower().endswith(".xlsx"):
        flash("Formato inválido. Envie um arquivo .xlsx.", "error")
        return redirect(url_for("procedimentos.pagina_base_procedimentos"))

    itens, erros = _ler_xlsx_procedimentos(f.read())

    if erros:
        flash(
            "Importação com problemas: " + " | ".join(erros[:6]) + (" ..." if len(erros) > 6 else ""),
            "error",
        )
        return redirect(url_for("procedimentos.pagina_base_procedimentos"))

    if not itens:
        flash("Nenhuma linha válida encontrada no Excel.", "info")
        return redirect(url_for("procedimentos.pagina_base_procedimentos"))

    conn = conectar_db()
    ensure_procedimentos_schema(conn)

    # PRAGMAs para performance (seguro pra import grande)
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("PRAGMA temp_store=MEMORY")
    except Exception:
        pass

    cur = conn.cursor()

    now = _now_sql()
    inseridos = 0
    atualizados = 0

    try:
        for it in itens:
            ins, upd = _upsert_regra(cur, it, now)
            inseridos += ins
            atualizados += upd

        conn.commit()
    except Exception as e:
        conn.rollback()
        flash(f"Falha ao importar: {e}", "error")
        return redirect(url_for("procedimentos.pagina_base_procedimentos"))

    flash(
        f"Importação concluída: {len(itens)} linha(s). Inseridos: {inseridos}. Atualizados: {atualizados}.",
        "success",
    )
    return redirect(url_for("procedimentos.pagina_base_procedimentos"))


@procedimentos_bp.get("/exportar.xlsx")
def exportar_procedimentos_excel():
    conn = conectar_db()
    ensure_procedimentos_schema(conn)
    cur = conn.cursor()

    cur.execute("""
        SELECT
          pa_cod, pa_cid, pa_descricao, pa_idademn, pa_idademx, pa_sexo, pa_cbo, pa_cbo_name
        FROM procedimentos
        WHERE ativo = 1
        ORDER BY pa_cbo ASC, pa_descricao COLLATE NOCASE ASC
    """)
    rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "PROCEDIMENTOS"

    ws.append(COLS_ESPERADAS)

    for r in rows:
        if hasattr(r, "keys"):
            ws.append(
                [
                    r["pa_cod"], r["pa_cid"], r["pa_descricao"], r["pa_idademn"],
                    r["pa_idademx"], r["pa_sexo"], r["pa_cbo"], r["pa_cbo_name"],
                ]
            )
        else:
            ws.append([r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7]])

    widths = {"A": 14, "B": 10, "C": 55, "D": 10, "E": 10, "F": 8, "G": 12, "H": 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    fname = f"procedimentos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@procedimentos_bp.get("/api/list")
def api_listar_procedimentos():
    """
    GET /procedimentos/api/list?page=1&per_page=20&cbo=&cid=&codigo=&nome=&idade=
    Retorna procedimentos do banco (paginação + filtros + contadores).
    """
    page = request.args.get("page", "1")
    per_page = request.args.get("per_page", "20")

    try:
        page_i = max(1, int(page))
    except Exception:
        page_i = 1

    try:
        per_i = int(per_page)
        if per_i < 1 or per_i > 200:
            per_i = 20
    except Exception:
        per_i = 20

    cbo    = (request.args.get("cbo") or "").strip()
    cid    = (request.args.get("cid") or "").strip()
    codigo = (request.args.get("codigo") or "").strip()
    nome   = (request.args.get("nome") or "").strip()

    idade_raw = (request.args.get("idade") or "").strip()
    idade = int(idade_raw) if idade_raw.isdigit() else None

    conn = conectar_db()
    ensure_procedimentos_schema(conn)
    cur = conn.cursor()

    # total geral (sem filtros)
    cur.execute("SELECT COUNT(1) FROM procedimentos WHERE ativo = 1")
    total_all = int(cur.fetchone()[0] or 0)

    where = ["ativo = 1"]
    params: list = []

    if cbo:
        where.append("pa_cbo = ?")
        params.append(cbo)

    if cid:
        where.append("(pa_cid = '' OR pa_cid = ?)")
        params.append(cid)

    if codigo:
        where.append("pa_cod LIKE ?")
        params.append(f"%{codigo}%")

    if nome:
        where.append("LOWER(pa_descricao) LIKE ?")
        params.append(f"%{nome.lower()}%")

    if idade is not None:
        where.append("(pa_idademn = -1 OR pa_idademn <= ?)")
        where.append("(pa_idademx = -1 OR pa_idademx >= ?)")
        params.extend([idade, idade])

    where_sql = " AND ".join(where)

    # total filtrado
    cur.execute(f"SELECT COUNT(1) FROM procedimentos WHERE {where_sql}", params)
    total_filtered = int(cur.fetchone()[0] or 0)

    pages = max(1, (total_filtered + per_i - 1) // per_i)
    if page_i > pages:
        page_i = pages

    offset = (page_i - 1) * per_i

    cur.execute(
        f"""
        SELECT pa_cod, pa_descricao, pa_cid, pa_cbo
          FROM procedimentos
         WHERE {where_sql}
         ORDER BY pa_cbo ASC, pa_descricao COLLATE NOCASE ASC
         LIMIT ? OFFSET ?
        """,
        params + [per_i, offset],
    )
    rows = cur.fetchall()

    items = [
        {"pa_cod": r[0], "pa_descricao": r[1], "pa_cid": r[2], "pa_cbo": r[3]}
        for r in rows
    ]

    return jsonify(
        {
            "ok": True,
            "page": page_i,
            "per_page": per_i,
            "total_all": total_all,
            "total_filtered": total_filtered,
            "pages": pages,
            "items": items,
        }
    )
