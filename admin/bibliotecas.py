# admin/bibliotecas.py
from __future__ import annotations

import io
import csv
from openpyxl import load_workbook
from flask import render_template, request, redirect, url_for, flash

from . import admin_bp, admin_required
from db import conectar_db

try:
    from .modulos import require_permission
except Exception:
    def require_permission(modulo_codigo: str, acao: str = "ver"):
        def deco(fn):
            return fn
        return deco

try:
    from log import registrar_log, log_erro
except Exception:
    def registrar_log(*args, **kwargs): pass
    def log_erro(*args, **kwargs): pass


# ============================================================
# SCHEMA POSTGRES
# ============================================================

def ensure_bibliotecas_postgres():
    conn = conectar_db()

    try:
        try:
            conn.rollback()
        except Exception:
            pass

        cur = conn.cursor()

        cur.execute("""
            CREATE TABLE IF NOT EXISTS ocupacoes (
                id SERIAL PRIMARY KEY,
                co_ocupacao VARCHAR(10) UNIQUE,
                no_ocupacao TEXT
            );
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS cid_catalogo (
                id SERIAL PRIMARY KEY,
                co_cid VARCHAR(10) UNIQUE,
                no_cid TEXT
            );
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS cep_ibge (
                id SERIAL PRIMARY KEY,
                cep VARCHAR(10),
                ibge VARCHAR(20),
                municipio TEXT,
                coduf VARCHAR(5),
                codmunicip VARCHAR(10),
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)

        cur.execute("CREATE INDEX IF NOT EXISTS idx_ocupacoes_codigo ON ocupacoes(co_ocupacao);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_ocupacoes_nome ON ocupacoes(no_ocupacao);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_cid_codigo ON cid_catalogo(co_cid);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_cid_nome ON cid_catalogo(no_cid);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_cep ON cep_ibge(cep);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_cep_ibge_municipio ON cep_ibge(municipio);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_cep_ibge_ibge ON cep_ibge(ibge);")

        conn.commit()
        cur.close()

    except Exception:
        conn.rollback()
        raise

    finally:
        conn.close()


# ============================================================
# HELPERS
# ============================================================

def _val(row, key: str, index: int = 0, default=None):
    if not row:
        return default

    if isinstance(row, dict):
        return row.get(key, default)

    try:
        return row[index]
    except Exception:
        return default


def _header_map(headers):
    mapa = {}

    for idx, h in enumerate(headers):
        chave = str(h or "").strip().lower()
        if chave:
            mapa[chave] = idx

    return mapa


def _normalizar_codigo_cbo(valor) -> str:
    codigo = str(valor or "").strip()

    if codigo.endswith(".0"):
        codigo = codigo[:-2]

    codigo = "".join(ch for ch in codigo if ch.isdigit())

    if not codigo:
        return ""

    return codigo.zfill(6)


def _normalizar_codigo_cid(valor) -> str:
    codigo = str(valor or "").strip().upper()
    codigo = codigo.replace(".", "").replace("-", "").strip()
    return codigo


def _pick_col(hmap: dict, nomes: list[str]):
    for nome in nomes:
        if nome in hmap:
            return hmap[nome]
    return None


def _filename(file_storage):
    return getattr(file_storage, "filename", "") or ""


# ============================================================
# IMPORTAÇÃO CBO
# ============================================================

def importar_cbo_xlsx(file_storage):
    ensure_bibliotecas_postgres()

    wb = load_workbook(file_storage, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)

    if not headers:
        raise ValueError("Arquivo vazio.")

    hmap = _header_map(headers)

    if "co_ocupacao" not in hmap or "no_ocupacao" not in hmap:
        raise ValueError("Colunas obrigatórias: co_ocupacao, no_ocupacao.")

    conn = conectar_db()

    try:
        try:
            conn.rollback()
        except Exception:
            pass

        cur = conn.cursor()

        processados = 0
        ignorados = 0
        lote = []

        for row in rows:
            codigo = _normalizar_codigo_cbo(row[hmap["co_ocupacao"]])
            nome = str(row[hmap["no_ocupacao"]] or "").strip()

            if not codigo or not nome:
                ignorados += 1
                continue

            lote.append((codigo, nome))

            if len(lote) >= 1000:
                cur.executemany("""
                    INSERT INTO ocupacoes (co_ocupacao, no_ocupacao)
                    VALUES (%s, %s)
                    ON CONFLICT (co_ocupacao)
                    DO UPDATE SET no_ocupacao = EXCLUDED.no_ocupacao;
                """, lote)

                processados += len(lote)
                lote.clear()

        if lote:
            cur.executemany("""
                INSERT INTO ocupacoes (co_ocupacao, no_ocupacao)
                VALUES (%s, %s)
                ON CONFLICT (co_ocupacao)
                DO UPDATE SET no_ocupacao = EXCLUDED.no_ocupacao;
            """, lote)

            processados += len(lote)

        conn.commit()
        cur.close()

        return processados, ignorados

    except Exception:
        conn.rollback()
        raise

    finally:
        conn.close()


# ============================================================
# IMPORTAÇÃO CID
# ============================================================

def importar_cid_xlsx(file_storage):
    ensure_bibliotecas_postgres()

    wb = load_workbook(file_storage, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)

    if not headers:
        raise ValueError("Arquivo vazio.")

    hmap = _header_map(headers)

    col_codigo = _pick_col(hmap, ["co_cid", "codigo", "código", "cid"])
    col_nome = _pick_col(hmap, ["no_cid", "descricao", "descrição", "nome"])

    if col_codigo is None or col_nome is None:
        raise ValueError(
            f"Colunas obrigatórias não encontradas. Cabeçalhos lidos: {list(hmap.keys())}"
        )

    conn = conectar_db()

    try:
        try:
            conn.rollback()
        except Exception:
            pass

        cur = conn.cursor()

        processados = 0
        ignorados = 0
        lote = []

        for row in rows:
            codigo = _normalizar_codigo_cid(row[col_codigo])
            nome = str(row[col_nome] or "").strip()

            if not codigo or not nome:
                ignorados += 1
                continue

            lote.append((codigo, nome))

            if len(lote) >= 1000:
                cur.executemany("""
                    INSERT INTO cid_catalogo (co_cid, no_cid)
                    VALUES (%s, %s)
                    ON CONFLICT (co_cid)
                    DO UPDATE SET no_cid = EXCLUDED.no_cid;
                """, lote)

                processados += len(lote)
                lote.clear()

        if lote:
            cur.executemany("""
                INSERT INTO cid_catalogo (co_cid, no_cid)
                VALUES (%s, %s)
                ON CONFLICT (co_cid)
                DO UPDATE SET no_cid = EXCLUDED.no_cid;
            """, lote)

            processados += len(lote)

        conn.commit()
        cur.close()

        return processados, ignorados

    except Exception:
        conn.rollback()
        raise

    finally:
        conn.close()


# ============================================================
# IMPORTAÇÃO CEP / IBGE
# ============================================================

def importar_cep_ibge_txt(file_storage, chunk_size=50000):
    ensure_bibliotecas_postgres()

    conn = conectar_db()
    processados = 0
    ignorados = 0

    try:
        try:
            conn.rollback()
        except Exception:
            pass

        cur = conn.cursor()

        cur.execute("TRUNCATE TABLE cep_ibge RESTART IDENTITY;")
        conn.commit()

        stream = io.TextIOWrapper(file_storage.stream, encoding="utf-8-sig", newline="")
        reader = csv.DictReader(stream, delimiter=";")

        lote = []

        for row in reader:
            row_norm = {
                str(k or "").strip().upper(): str(v or "").strip()
                for k, v in row.items()
            }

            cep = row_norm.get("CEP", "")
            ibge = row_norm.get("IBGE", "")
            municipio = row_norm.get("MUNICIPIO", "")
            coduf = row_norm.get("CODUF", "")
            codmunicip = row_norm.get("CODMUNIC", "")
            criado_em = row_norm.get("CRIADO_EM") or None

            if not cep or not ibge or not municipio:
                ignorados += 1
                continue

            lote.append((cep, ibge, municipio, coduf, codmunicip, criado_em))

            if len(lote) >= chunk_size:
                cur.executemany("""
                    INSERT INTO cep_ibge
                    (cep, ibge, municipio, coduf, codmunicip, criado_em)
                    VALUES (%s, %s, %s, %s, %s, COALESCE(%s::timestamp, CURRENT_TIMESTAMP));
                """, lote)

                conn.commit()
                processados += len(lote)
                lote.clear()

        if lote:
            cur.executemany("""
                INSERT INTO cep_ibge
                (cep, ibge, municipio, coduf, codmunicip, criado_em)
                VALUES (%s, %s, %s, %s, %s, COALESCE(%s::timestamp, CURRENT_TIMESTAMP));
            """, lote)

            conn.commit()
            processados += len(lote)

        cur.close()

        return processados, ignorados

    except Exception:
        conn.rollback()
        raise

    finally:
        conn.close()


# ============================================================
# ROTA CBO
# ============================================================

@admin_bp.route("/cbo", methods=["GET", "POST"])
@admin_required
@require_permission("admin_cbo", "ver")
def biblioteca_cbo():
    ensure_bibliotecas_postgres()

    if request.method == "POST":
        arquivo = request.files.get("arquivo")

        if not arquivo or arquivo.filename == "":
            flash("Selecione um arquivo XLSX.", "error")
            return redirect(url_for("admin.biblioteca_cbo"))

        try:
            p, i = importar_cbo_xlsx(arquivo)

            registrar_log(
                modulo="admin_cbo",
                acao="importar",
                entidade="ocupacoes",
                descricao="Importou biblioteca CBO.",
                detalhes={
                    "arquivo": _filename(arquivo),
                    "processados": p,
                    "ignorados": i,
                },
            )

            flash(f"CBO importado com sucesso: {p} registros. Ignorados: {i}.", "success")

        except Exception as e:
            log_erro(
                "admin_cbo",
                e,
                entidade="ocupacoes",
                descricao="Erro ao importar CBO.",
                detalhes={"arquivo": _filename(arquivo)},
            )
            flash(f"Erro ao importar CBO: {e}", "error")

        return redirect(url_for("admin.biblioteca_cbo"))

    q = (request.args.get("q") or "").strip()
    itens = []

    conn = conectar_db()

    try:
        try:
            conn.rollback()
        except Exception:
            pass

        cur = conn.cursor()

        if q:
            like = f"%{q}%"
            cur.execute("""
                SELECT
                    co_ocupacao AS codigo,
                    no_ocupacao AS descricao
                FROM ocupacoes
                WHERE co_ocupacao ILIKE %s
                   OR no_ocupacao ILIKE %s
                ORDER BY co_ocupacao
                LIMIT 500;
            """, (like, like))
        else:
            cur.execute("""
                SELECT
                    co_ocupacao AS codigo,
                    no_ocupacao AS descricao
                FROM ocupacoes
                ORDER BY co_ocupacao
                LIMIT 500;
            """)

        rows = cur.fetchall() or []

        for row in rows:
            itens.append({
                "codigo": _val(row, "codigo", 0, ""),
                "descricao": _val(row, "descricao", 1, ""),
            })

        cur.close()

        registrar_log(
            modulo="admin_cbo",
            acao="visualizar",
            entidade="ocupacoes",
            descricao="Visualizou biblioteca CBO.",
            detalhes={"q": q, "total_exibido": len(itens)},
        )

    except Exception as e:
        conn.rollback()
        log_erro("admin_cbo", e, entidade="ocupacoes", descricao="Erro ao carregar CBOs.", detalhes={"q": q})
        flash(f"Erro ao carregar CBOs: {e}", "error")

    finally:
        conn.close()

    return render_template("cbo.html", titulo="Biblioteca CBO", q=q, itens=itens)


# ============================================================
# ROTA CID
# ============================================================

@admin_bp.route("/cid", methods=["GET", "POST"])
@admin_required
@require_permission("admin_cid", "ver")
def biblioteca_cid():
    ensure_bibliotecas_postgres()

    if request.method == "POST":
        arquivo = request.files.get("arquivo")

        if not arquivo or arquivo.filename == "":
            flash("Selecione um arquivo XLSX.", "error")
            return redirect(url_for("admin.biblioteca_cid"))

        try:
            p, i = importar_cid_xlsx(arquivo)

            registrar_log(
                modulo="admin_cid",
                acao="importar",
                entidade="cid_catalogo",
                descricao="Importou biblioteca CID.",
                detalhes={
                    "arquivo": _filename(arquivo),
                    "processados": p,
                    "ignorados": i,
                },
            )

            flash(f"CID importado com sucesso: {p} registros. Ignorados: {i}.", "success")

        except Exception as e:
            log_erro(
                "admin_cid",
                e,
                entidade="cid_catalogo",
                descricao="Erro ao importar CID.",
                detalhes={"arquivo": _filename(arquivo)},
            )
            flash(f"Erro ao importar CID: {e}", "error")

        return redirect(url_for("admin.biblioteca_cid"))

    q = (request.args.get("q") or "").strip()
    pagina = request.args.get("pagina", 1, type=int)

    por_pagina = 50
    pagina = max(pagina, 1)
    offset = (pagina - 1) * por_pagina

    itens = []
    total = 0

    conn = conectar_db()

    try:
        try:
            conn.rollback()
        except Exception:
            pass

        cur = conn.cursor()

        if q:
            like = f"%{q}%"

            cur.execute("""
                SELECT COUNT(*) AS total
                FROM cid_catalogo
                WHERE co_cid ILIKE %s
                   OR no_cid ILIKE %s;
            """, (like, like))

            total = int(_val(cur.fetchone(), "total", 0, 0) or 0)

            cur.execute("""
                SELECT
                    co_cid AS codigo,
                    no_cid AS descricao
                FROM cid_catalogo
                WHERE co_cid ILIKE %s
                   OR no_cid ILIKE %s
                ORDER BY co_cid
                LIMIT %s OFFSET %s;
            """, (like, like, por_pagina, offset))

        else:
            cur.execute("SELECT COUNT(*) AS total FROM cid_catalogo;")
            total = int(_val(cur.fetchone(), "total", 0, 0) or 0)

            cur.execute("""
                SELECT
                    co_cid AS codigo,
                    no_cid AS descricao
                FROM cid_catalogo
                ORDER BY co_cid
                LIMIT %s OFFSET %s;
            """, (por_pagina, offset))

        rows = cur.fetchall() or []

        for row in rows:
            itens.append({
                "codigo": _val(row, "codigo", 0, ""),
                "descricao": _val(row, "descricao", 1, ""),
            })

        cur.close()

        registrar_log(
            modulo="admin_cid",
            acao="visualizar",
            entidade="cid_catalogo",
            descricao="Visualizou biblioteca CID.",
            detalhes={
                "q": q,
                "pagina": pagina,
                "total": total,
                "total_exibido": len(itens),
            },
        )

    except Exception as e:
        conn.rollback()
        log_erro("admin_cid", e, entidade="cid_catalogo", descricao="Erro ao carregar CIDs.", detalhes={"q": q})
        flash(f"Erro ao carregar CIDs: {e}", "error")

    finally:
        conn.close()

    total_paginas = max((total + por_pagina - 1) // por_pagina, 1)

    if pagina > total_paginas:
        pagina = total_paginas

    return render_template(
        "cid.html",
        titulo="Biblioteca CID",
        q=q,
        itens=itens,
        pagina=pagina,
        total_paginas=total_paginas,
        total=total,
        por_pagina=por_pagina,
    )


# ============================================================
# ROTA CEP / IBGE
# ============================================================

@admin_bp.route("/cep-ibge", methods=["GET", "POST"])
@admin_required
@require_permission("admin_cep_ibge", "ver")
def biblioteca_cep_ibge():
    ensure_bibliotecas_postgres()

    if request.method == "POST":
        arquivo = request.files.get("arquivo")

        if not arquivo or arquivo.filename == "":
            flash("Selecione um arquivo TXT separado por ponto e vírgula.", "error")
            return redirect(url_for("admin.biblioteca_cep_ibge"))

        try:
            p, i = importar_cep_ibge_txt(arquivo)

            registrar_log(
                modulo="admin_cep_ibge",
                acao="importar",
                entidade="cep_ibge",
                descricao="Importou biblioteca CEP/IBGE.",
                detalhes={
                    "arquivo": _filename(arquivo),
                    "processados": p,
                    "ignorados": i,
                },
            )

            flash(f"CEP/IBGE importado com sucesso: {p} registros. Ignorados: {i}.", "success")

        except Exception as e:
            log_erro(
                "admin_cep_ibge",
                e,
                entidade="cep_ibge",
                descricao="Erro ao importar CEP/IBGE.",
                detalhes={"arquivo": _filename(arquivo)},
            )
            flash(f"Erro ao importar CEP/IBGE: {e}", "error")

        return redirect(url_for("admin.biblioteca_cep_ibge"))

    q = (request.args.get("q") or "").strip()
    pagina = request.args.get("pagina", 1, type=int)

    por_pagina = 50
    pagina = max(pagina, 1)
    offset = (pagina - 1) * por_pagina

    itens = []
    total = 0

    conn = conectar_db()

    try:
        try:
            conn.rollback()
        except Exception:
            pass

        cur = conn.cursor()

        if q:
            like = f"%{q}%"

            cur.execute("""
                SELECT COUNT(*) AS total
                FROM cep_ibge
                WHERE cep ILIKE %s
                   OR ibge ILIKE %s
                   OR municipio ILIKE %s
                   OR coduf ILIKE %s
                   OR codmunicip ILIKE %s;
            """, (like, like, like, like, like))

            total = int(_val(cur.fetchone(), "total", 0, 0) or 0)

            cur.execute("""
                SELECT
                    cep,
                    ibge,
                    municipio,
                    coduf,
                    codmunicip,
                    criado_em
                FROM cep_ibge
                WHERE cep ILIKE %s
                   OR ibge ILIKE %s
                   OR municipio ILIKE %s
                   OR coduf ILIKE %s
                   OR codmunicip ILIKE %s
                ORDER BY municipio, cep
                LIMIT %s OFFSET %s;
            """, (like, like, like, like, like, por_pagina, offset))

        else:
            cur.execute("SELECT COUNT(*) AS total FROM cep_ibge;")
            total = int(_val(cur.fetchone(), "total", 0, 0) or 0)

            cur.execute("""
                SELECT
                    cep,
                    ibge,
                    municipio,
                    coduf,
                    codmunicip,
                    criado_em
                FROM cep_ibge
                ORDER BY municipio, cep
                LIMIT %s OFFSET %s;
            """, (por_pagina, offset))

        rows = cur.fetchall() or []

        for row in rows:
            itens.append({
                "cep": _val(row, "cep", 0, ""),
                "ibge": _val(row, "ibge", 1, ""),
                "municipio": _val(row, "municipio", 2, ""),
                "coduf": _val(row, "coduf", 3, ""),
                "codmunicip": _val(row, "codmunicip", 4, ""),
                "criado_em": _val(row, "criado_em", 5, ""),
            })

        cur.close()

        registrar_log(
            modulo="admin_cep_ibge",
            acao="visualizar",
            entidade="cep_ibge",
            descricao="Visualizou biblioteca CEP/IBGE.",
            detalhes={
                "q": q,
                "pagina": pagina,
                "total": total,
                "total_exibido": len(itens),
            },
        )

    except Exception as e:
        conn.rollback()
        log_erro("admin_cep_ibge", e, entidade="cep_ibge", descricao="Erro ao carregar CEP/IBGE.", detalhes={"q": q})
        flash(f"Erro ao carregar CEP/IBGE: {e}", "error")

    finally:
        conn.close()

    total_paginas = max((total + por_pagina - 1) // por_pagina, 1)

    if pagina > total_paginas:
        pagina = total_paginas

    return render_template(
        "cep_ibge.html",
        titulo="Biblioteca CEP/IBGE",
        q=q,
        itens=itens,
        pagina=pagina,
        total_paginas=total_paginas,
        total=total,
        por_pagina=por_pagina,
    )