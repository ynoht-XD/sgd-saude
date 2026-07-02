# -*- coding: utf-8 -*-
from __future__ import annotations

import io
import re
import traceback
from datetime import datetime, date
from typing import Any

from flask import request, send_file, flash, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from db import conectar_db
from . import export_bp

try:
    from auth import login_required
except ImportError:
    def login_required(f):
        return f


_RE_DIGITS = re.compile(r"\D+")


def _digits(v: Any) -> str:
    if v is None:
        return ""
    return _RE_DIGITS.sub("", str(v))


def _zfill(v: Any, n: int) -> str:
    return _digits(v).zfill(n)[:n]


def _ljust(v: Any, n: int) -> str:
    v = "" if v is None else str(v)
    return v[:n].ljust(n)


def _safe_digits(value: Any) -> str:
    return _digits(str(value or "").strip())


def _safe_zfill(value: Any, size: int, default: str = "") -> str:
    raw = _safe_digits(value)
    if not raw:
        raw = _safe_digits(default)
    return str(raw).zfill(size)[:size]


def _safe_ljust_text(value: Any, size: int, default: str = "") -> str:
    txt = str(value or default or "").strip()
    return txt[:size].ljust(size)


def _fmt_date_yyyymmdd(dt) -> str:
    if not dt:
        return "00000000"
    if isinstance(dt, str):
        s = dt.strip()
        if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
            return s.replace("-", "")
        if re.match(r"^\d{2}/\d{2}/\d{4}$", s):
            d, m, y = s.split("/")
            return f"{y}{m}{d}"
        return "00000000"
    return dt.strftime("%Y%m%d")


def _norm_comp_in(v: str | None) -> str:
    """
    Retorna MMYYYY.
    Aceita: MM/YYYY, MMYYYY, YYYYMM.
    """
    if not v:
        return ""
    d = _digits(v)
    if len(d) != 6:
        return ""
    if int(d[:2]) > 12:
        return d[4:6] + d[0:4]
    return d


def _comp_mmyyyy_to_yyyymm(mmyyyy: str) -> str:
    if not mmyyyy or len(mmyyyy) != 6:
        return "000000"
    return mmyyyy[2:6] + mmyyyy[0:2]


def _comp_extensao(mmyyyy: str) -> str:
    meses = {
        "01": "JAN",
        "02": "FEV",
        "03": "MAR",
        "04": "ABR",
        "05": "MAI",
        "06": "JUN",
        "07": "JUL",
        "08": "AGO",
        "09": "SET",
        "10": "OUT",
        "11": "NOV",
        "12": "DEZ",
    }
    mm = (mmyyyy or "")[:2]
    return meses.get(mm, "TXT")


def _is_numeric_like(value: Any) -> bool:
    s = str(value or "").strip()
    if not s:
        return False
    if s.endswith(".0"):
        s = s[:-2]
    return s.isdigit()


def _table_exists(cur, table_name: str) -> bool:
    cur.execute(
        """
        SELECT EXISTS (
            SELECT 1
              FROM information_schema.tables
             WHERE table_schema = 'public'
               AND table_name = %s
        ) AS existe
        """,
        (table_name,),
    )
    row = cur.fetchone()
    return bool(row["existe"]) if row else False


def _columns(cur, table_name: str) -> set[str]:
    cur.execute(
        """
        SELECT column_name
          FROM information_schema.columns
         WHERE table_schema = 'public'
           AND table_name = %s
        """,
        (table_name,),
    )
    return {r["column_name"] for r in cur.fetchall()}


def _select_existing(cur, desired: list[str], table_name: str = "apac") -> list[str]:
    existentes = _columns(cur, table_name)
    return [c for c in desired if c in existentes]


def _build_apac_where(args, existentes: set[str]) -> tuple[str, list]:
    where = ["1=1"]
    params = []

    comp_in = (args.get("competencia") or "").strip()
    comp = _norm_comp_in(comp_in)

    if comp and "competencia" in existentes:
        where.append("REPLACE(COALESCE(competencia::text,''), '/', '') = %s")
        params.append(comp)

    nome = (args.get("nome") or "").strip()
    if nome:
        if "nome_paciente" in existentes:
            where.append("LOWER(COALESCE(nome_paciente,'')) LIKE LOWER(%s)")
            params.append(f"%{nome}%")
        elif "nome" in existentes:
            where.append("LOWER(COALESCE(nome,'')) LIKE LOWER(%s)")
            params.append(f"%{nome}%")

    cep = _digits(args.get("cep") or "")
    if cep and "cep" in existentes:
        where.append("regexp_replace(COALESCE(cep::text,''), '\\D', '', 'g') LIKE %s")
        params.append(f"%{cep}%")

    numero_apac = _digits(args.get("numero_apac") or "")
    if numero_apac and "numero_apac" in existentes:
        where.append("regexp_replace(COALESCE(numero_apac::text,''), '\\D', '', 'g') LIKE %s")
        params.append(f"%{numero_apac}%")

    status = (args.get("status") or "").strip()
    if status and "status" in existentes:
        where.append("COALESCE(status,'') = %s")
        params.append(status)

    status_entrega = (args.get("status_entrega") or "").strip()
    if status_entrega and "status_entrega" in existentes:
        where.append("COALESCE(status_entrega,'') = %s")
        params.append(status_entrega)

    nota_fiscal = (args.get("nota_fiscal") or "").strip()
    if nota_fiscal and "nota_fiscal" in existentes:
        where.append("COALESCE(nota_fiscal::text,'') ILIKE %s")
        params.append(f"%{nota_fiscal}%")

    competencia_nota = (args.get("competencia_nota") or "").strip()
    if competencia_nota and "competencia_nota" in existentes:
        where.append("COALESCE(competencia_nota::text,'') ILIKE %s")
        params.append(f"%{competencia_nota}%")

    fornecedor = (args.get("fornecedor") or "").strip()
    if fornecedor and "fornecedor" in existentes:
        where.append("COALESCE(fornecedor,'') ILIKE %s")
        params.append(f"%{fornecedor}%")

    local_entrega = (args.get("local_entrega") or "").strip()
    if local_entrega and "local_entrega" in existentes:
        where.append("COALESCE(local_entrega,'') ILIKE %s")
        params.append(f"%{local_entrega}%")

    processado = (args.get("processado") or "").strip().lower()
    if "processado" in existentes:
        if processado in ("true", "t", "1", "yes", "y", "sim", "s"):
            where.append("processado = TRUE")
        elif processado in ("false", "f", "0", "no", "n", "nao", "não"):
            where.append("processado = FALSE")

    bpai = (args.get("bpai") or "").strip().lower()
    if "bpai" in existentes:
        if bpai in ("true", "t", "1", "yes", "y", "sim", "s"):
            where.append("bpai = TRUE")
        elif bpai in ("false", "f", "0", "no", "n", "nao", "não"):
            where.append("bpai = FALSE")

    return " AND ".join(where), params


def _competencia_mais_recente(cur, where_sql: str = "1=1", params: list | None = None) -> str:
    params = params or []
    try:
        cur.execute(
            f"""
            SELECT REPLACE(COALESCE(competencia::text,''), '/', '') AS comp
              FROM public.apac
             WHERE {where_sql}
               AND REPLACE(COALESCE(competencia::text,''), '/', '') ~ '^[0-9]{{6}}$'
             ORDER BY id DESC
             LIMIT 1
            """,
            params,
        )
        row = cur.fetchone()
        return _norm_comp_in(row["comp"] if row else "") or "000000"
    except Exception:
        return "000000"


@export_bp.get("/apac/excel", endpoint="apacs_excel")
@login_required
def apacs_excel():
    conn = conectar_db()
    cur = conn.cursor()

    try:
        if not _table_exists(cur, "apac"):
            flash("Tabela APAC ainda não existe.", "warning")
            return redirect(url_for("export.apac_view"))

        existentes = _columns(cur, "apac")
        where_sql, params = _build_apac_where(request.args, existentes)

        select_cols = _select_existing(cur, [
            "id", "numero_apac", "competencia", "nome_paciente", "nome",
            "prontuario", "procedimento", "codigo_procedimento", "quantidade",
            "cnes", "tipo_apac", "nacionalidade", "data_nascimento", "nascimento",
            "sexo", "raca", "cep", "endereco", "logradouro", "rua", "numero",
            "numero_casa", "bairro", "cns", "cpf", "nome_mae", "mae",
            "responsavel", "servico", "classificacao", "cid", "cid2",
            "descricao_diagnostico", "carater_atendimento",
            "nome_solicitante", "cns_solicitante", "data_solicitacao",
            "nome_autorizador", "cns_autorizador", "data_autorizacao",
            "orgao_emissor", "data_inicial", "data_final",
            "status", "nota_fiscal", "data_nota_fiscal", "data_entrada_nf",
            "competencia_nota", "protocolo_nota", "fornecedor", "local_entrega",
            "status_entrega", "data_pedido", "data_entrega",
            "obs_nota", "obs_pedido", "obs_entrega", "obs_geral",
            "processado", "bpai", "sms_enviado", "criado_em", "atualizado_em",
        ])

        if not select_cols:
            flash("Nenhuma coluna APAC encontrada para exportar.", "warning")
            return redirect(url_for("export.apac_view"))

        sql = f"""
            SELECT {", ".join(select_cols)}
              FROM public.apac
             WHERE {where_sql}
             ORDER BY id DESC
        """

        cur.execute(sql, params)
        rows = [dict(r) for r in cur.fetchall()]

    finally:
        cur.close()
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "APACs"

    ws.append(select_cols)

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E85")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for col_idx in range(1, len(select_cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for row in rows:
        ws.append([row.get(c) for c in select_cols])

    date_cols = {
        "data_nascimento", "nascimento", "data_solicitacao", "data_autorizacao",
        "data_inicial", "data_final", "data_nota_fiscal", "data_pedido",
        "data_entrega", "data_entrada_nf", "criado_em", "atualizado_em",
    }

    col_index_name = {i + 1: name for i, name in enumerate(select_cols)}

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if col_index_name[c] in date_cols:
                val = ws.cell(row=r, column=c).value
                if isinstance(val, (datetime, date)):
                    ws.cell(row=r, column=c).number_format = "DD/MM/YYYY"

    ws.freeze_panes = "A2"

    for c in range(1, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=c).column_letter
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                max_len = max(max_len, min(len(str(v)), 60))
        ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 64))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    comp = _norm_comp_in(request.args.get("competencia"))
    sufixo = f"_{comp[:2]}-{comp[2:]}" if comp else ""

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"apacs_exportadas{sufixo}.xlsx",
    )


@export_bp.get("/apac/txt", endpoint="apacs_txt")
@login_required
def apacs_txt():
    conn = conectar_db()
    cur = conn.cursor()

    try:
        if not _table_exists(cur, "apac"):
            flash("Tabela APAC ainda não existe.", "warning")
            return redirect(url_for("export.apac_view"))

        existentes = _columns(cur, "apac")
        where_sql, params = _build_apac_where(request.args, existentes)

        comp_forcada = _norm_comp_in(request.args.get("competencia"))
        mmyyyy_header = comp_forcada or _competencia_mais_recente(cur, where_sql, params)
        yyyymm_header = _comp_mmyyyy_to_yyyymm(mmyyyy_header)

        cur.execute(f"SELECT COUNT(*) AS total FROM public.apac WHERE {where_sql}", params)
        qtd = int(cur.fetchone()["total"] or 0)

        cur.execute(
            f"""
            SELECT
                COALESCE(SUM(CASE
                    WHEN regexp_replace(COALESCE(codigo_procedimento::text,''), '\\D', '', 'g') ~ '^[0-9]+$'
                    THEN regexp_replace(COALESCE(codigo_procedimento::text,''), '\\D', '', 'g')::bigint
                    ELSE 0 END), 0) AS soma_codigos,
                COALESCE(SUM(CASE
                    WHEN regexp_replace(COALESCE(quantidade::text,''), '\\D', '', 'g') ~ '^[0-9]+$'
                    THEN regexp_replace(COALESCE(quantidade::text,''), '\\D', '', 'g')::bigint
                    ELSE 0 END), 0) AS soma_qtd,
                COALESCE(SUM(CASE
                    WHEN regexp_replace(COALESCE(numero_apac::text,''), '\\D', '', 'g') ~ '^[0-9]+$'
                    THEN regexp_replace(COALESCE(numero_apac::text,''), '\\D', '', 'g')::bigint
                    ELSE 0 END), 0) AS soma_apac
              FROM public.apac
             WHERE {where_sql}
            """,
            params,
        )

        soma = cur.fetchone()
        total_controle = int(soma["soma_codigos"] or 0) + int(soma["soma_qtd"] or 0) + int(soma["soma_apac"] or 0)
        campo_controle = (total_controle % 1111) + 1111

        cabecalho = (
            f"01#APAC{yyyymm_header}{str(qtd).zfill(6)}{campo_controle}"
            f"{_ljust('ACRESC', 30)}"
            f"{_ljust('ACRESC', 6)}"
            f"{_zfill('09553609000162', 14)}"
            f"{_ljust('SECRETARIA MUNICIPAL DE SAUDE DE PENEDO', 40)}"
            f"M"
            f"{datetime.now().strftime('%Y%m%d')}"
            f"{_ljust('Versao 03.14', 15)}"
        )

        cur.execute(
            f"""
            SELECT
                numero_apac,
                REPLACE(COALESCE(competencia::text,''), '/', '') AS comp,
                data_inicial,
                data_final,
                tipo_apac,
                COALESCE(nome_paciente, nome, '') AS nome_paciente,
                COALESCE(nome_mae, mae, '') AS nome_mae,
                COALESCE(endereco, logradouro, rua, '') AS endereco,
                COALESCE(numero, numero_casa, '') AS numero,
                cep,
                COALESCE(data_nascimento, nascimento) AS data_nascimento,
                sexo,
                nome_solicitante,
                codigo_procedimento,
                motivo_saida,
                data_alta,
                nome_autorizador,
                cns,
                cns_autorizador,
                prontuario,
                bairro,
                data_solicitacao,
                data_autorizacao,
                CAST(orgao_emissor AS TEXT) AS orgao_emissor,
                carater_atendimento,
                responsavel,
                cns_executante,
                cid,
                cbo_executante,
                quantidade,
                servico,
                classificacao,
                cns_solicitante
              FROM public.apac
             WHERE {where_sql}
             ORDER BY numero_apac
            """,
            params,
        )

        apacs = cur.fetchall()

        cep_para_ibge = {
            "57200000": "270670",
            "57210000": "270680",
            "57230000": "270230",
            "57290000": "270750",
            "57300005": "270030",
            "57280000": "270320",
            "57220000": "270270",
            "49980000": "280440",
            "57600000": "270880",
            "57380000": "270820",
        }

        linhas = []
        ignoradas = 0

        for idx, r in enumerate(apacs, start=1):
            try:
                numero_apac_raw = r["numero_apac"]
                codigo_proc_raw = r["codigo_procedimento"]

                if not _is_numeric_like(numero_apac_raw):
                    ignoradas += 1
                    print(f"⚠️ APAC ignorada #{idx}: numero_apac inválido -> {numero_apac_raw!r}")
                    continue

                if not _is_numeric_like(codigo_proc_raw):
                    ignoradas += 1
                    print(f"⚠️ APAC ignorada #{idx}: codigo_procedimento inválido -> {codigo_proc_raw!r}")
                    continue

                numero_apac_13 = _safe_zfill(numero_apac_raw, 13)
                cod_proc = _safe_zfill(codigo_proc_raw, 10)

                comp_apac_mmyyyy = _norm_comp_in(r["comp"]) or mmyyyy_header
                comp_apac_yyyymm = _comp_mmyyyy_to_yyyymm(comp_apac_mmyyyy)

                cep = _safe_zfill(r["cep"], 8)
                ibge = cep_para_ibge.get(cep, "").ljust(7)[:7]

                linha14 = (
                    f"14{comp_apac_yyyymm}{numero_apac_13}"
                    f"27"
                    f"6097367"
                    f"{comp_apac_yyyymm}01"
                    f"{_fmt_date_yyyymmdd(r['data_inicial'])}{_fmt_date_yyyymmdd(r['data_final'])}"
                    f"00"
                    f"{(str(r['tipo_apac'] or ' ').strip()[:1] or ' ').upper()}"
                    f"{_ljust(r['nome_paciente'], 30)}"
                    f"{_ljust(r['nome_mae'], 30)}"
                    f"{_ljust(r['endereco'], 30)}"
                    f"{_safe_ljust_text(_safe_digits(r['numero']), 5)}"
                    f"{' ' * 10}"
                    f"{cep}"
                    f"{ibge}"
                    f"{_fmt_date_yyyymmdd(r['data_nascimento'])}"
                    f"{(str(r['sexo'] or 'U').strip().upper()[:1]) or 'U'}"
                    f"{_ljust(r['nome_solicitante'], 30)}"
                    f"{cod_proc}"
                    f"{_safe_ljust_text(r['motivo_saida'], 2)}"
                    f"{_fmt_date_yyyymmdd(r['data_alta'])}"
                    f"{_ljust(r['nome_autorizador'], 30)}"
                    f"{_safe_zfill(r['cns'], 15)}"
                    f"{_safe_zfill(r['cns_solicitante'], 15)}"
                    f"{_safe_zfill(r['cns_autorizador'], 15)}"
                    f"{' ' * 4}"
                    f"{_safe_ljust_text(r['prontuario'], 10)}"
                    f"6097367"
                    f"{_fmt_date_yyyymmdd(r['data_solicitacao'])}"
                    f"{_fmt_date_yyyymmdd(r['data_autorizacao'])}"
                    f"{_safe_ljust_text(r['orgao_emissor'], 10, 'M27670301')}"
                    f"{_safe_zfill(r['carater_atendimento'], 2, '01')}"
                    f"{'0' * 13}"
                    f"03"
                    f"{_ljust(r['responsavel'] or r['nome_mae'], 30)}"
                    f"010"
                    f"{' ' * 4}"
                    f"081"
                    f"{_ljust(r['bairro'], 30)}"
                    f"{' ' * 11}"
                    f"{' ' * 40}"
                    f"{_safe_zfill(r['cns_executante'], 15)}"
                    f"{'0' * 11}"
                    f"{' ' * 9}"
                    f"{' '}"
                    f" "
                )

                linha06 = (
                    f"06{comp_apac_yyyymm}{numero_apac_13}"
                    f"{_safe_ljust_text(str(r['cid'] or '').upper(), 4)}"
                )

                linha13 = (
                    f"13{comp_apac_yyyymm}{numero_apac_13}"
                    f"{cod_proc}"
                    f"{_safe_ljust_text(r['cbo_executante'], 6)}"
                    f"{_safe_zfill(r['quantidade'], 7, '0')}"
                    f"{' ' * 14}"
                    f"{' ' * 6}"
                    f"{' ' * 4}"
                    f"{' ' * 4}"
                    f"{_safe_zfill(r['servico'], 3, '0')}"
                    f"{_safe_zfill(r['classificacao'], 3, '0')}"
                    f"{' ' * 8}"
                    f"{' ' * 4}"
                    f"{' ' * 7}"
                )

                linhas.extend([linha14, linha06, linha13])

            except Exception as e:
                ignoradas += 1
                print(f"❌ Erro montando APAC #{idx}: {e}")
                traceback.print_exc()

        if ignoradas:
            print(f"⚠️ Total de APACs ignoradas na exportação: {ignoradas}")

        conteudo = cabecalho.rstrip() + "\r\n" + "\r\n".join(linhas) + "\r\n" + "\x1A"

    finally:
        cur.close()
        conn.close()

    buffer = io.BytesIO(conteudo.encode("latin-1"))
    buffer.seek(0)

    extensao = _comp_extensao(mmyyyy_header)
    filename = f"AP{yyyymm_header}.{extensao}"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/octet-stream",
    )
